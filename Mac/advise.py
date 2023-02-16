import re #used for regular expressions
from datetime import datetime #use to get current date
import pandas as pd #used to structure data
import openpyxl #used to create excel files
from openpyxl.workbook import Workbook #used to generate worksheets in memory
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side #used for excel border styles
from openpyxl.styles import PatternFill #used for excell cell background colors
from openpyxl.styles.colors import Color
from openpyxl.styles import Alignment

#global variables for border styles
top = Border(left = Side(style = 'thin'),
              right = Side(style = 'thin'),
              top = Side(style = 'thick'),
              bottom = Side(style = 'thin'))import re  # used for regular expressions
from datetime import datetime  # use to get current date
import pandas as pd  # used to structure data
import openpyxl  # used to create excel files
from openpyxl.workbook import Workbook  # used to generate worksheets in memory
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side  # used for excel border styles
from openpyxl.styles import PatternFill  # used for excell cell background colors
from openpyxl.styles.colors import Color
from openpyxl.styles import Alignment

# global variables for border styles
top = Border(left=Side(style='thin'),
             right=Side(style='thin'),
             top=Side(style='thick'),
             bottom=Side(style='thin'))
top_left = Border(left=Side(style='thick'),
                  right=Side(style='thin'),
                  top=Side(style='thick'),
                  bottom=Side(style='thin'))
top_right = Border(left=Side(style='thin'),
                   right=Side(style='thick'),
                   top=Side(style='thick'),
                   bottom=Side(style='thin'))
right = Border(left=Side(style='thin'),
               right=Side(style='thick'),
               top=Side(style='thin'),
               bottom=Side(style='thin'))
bottom_right = Border(left=Side(style='thin'),
                      right=Side(style='thick'),
                      top=Side(style='thin'),
                      bottom=Side(style='thick'))
bottom = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thick'))
bottom_left = Border(left=Side(style='thick'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thick'))
left = Border(left=Side(style='thick'),
              right=Side(style='thin'),
              top=Side(style='thin'),
              bottom=Side(style='thin'))
other = Border(left=Side(style='thin'),
               right=Side(style='thin'),
               top=Side(style='thin'),
               bottom=Side(style='thin'))
# global variables for cell background colors
redFill = PatternFill('solid', fgColor='FF0000')
greenFill = PatternFill('solid', fgColor='00FF00')
yellowFill = PatternFill('solid', fgColor='FFFF00')
blueFill = PatternFill('solid', fgColor='00BFFF')


# at least 8 csci courses, 60 credits, and csci 287 before 400-level csci courses
# method to check if a course has a passing grade where a passing grade is 'C'
def passing_grade(grade, sem, curr_sem):
    if grade <= "C" or grade == "S" or grade == "TR" or grade == "In progress" \
            or grade == "SP":
        # add a check for future courses, which should be omitted
        # from course recommendations
        # If course is from this semester, include it
        if grade == "In progress":
            if sem == curr_sem:
                return True
            # If course is from future semester, omit it
            else:
                return False
        # If course is not from this semester, then it has already
        # been taken
        else:
            return True
    # If we reached this, then we reached a failing course somehow
    else:
        return False


# method to check if a course has a passing grade where a passing grade is 'D'
def passing_grade_alt(grade, sem, curr_sem):
    if grade <= "D" or grade == "S" or grade == "TR" or grade == "In progress" \
            or grade == "SP":
        if grade == "In progress":
            if sem == curr_sem:
                return True
            else:
                return False
        else:
            return True
    else:
        return False


# Check for CSCI Elective
def csci_elec(course):
    # 'CSCI 312', 'CSCI 361', 'CSCI 389', 'CSCI 396', 'CSCI 402', 'CSCI 450'
    # 'CSCI 451', 'CSCI 453', 'CSCI 495', 'CSCI 496'
    elec = ['CSCI 312', 'CSCI 361', 'CSCI 389', 'CSCI 396', 'CSCI 402', \
            'CSCI 450', 'CSCI 451', 'CSCI 453', 'CSCI 495', 'CSCI 496']
    if course in elec:
        return True
    else:
        return False


# Check for Global Studies Elective
def gs_elec(course):
    # 'SPAN', 'FREN', 'GERM', 'HIST 114', 'HIST 115', 'ARTS 199'
    elec = ['SPAN', 'FREN', 'GERM', 'HIST 114', 'HIST 115', 'ARTS 199']

    for i in elec:
        if re.search(i, course):
            return True
    return False


# Check for Social Sciences Elective
def ss_elec(course):
    # 'SOCI', 'PSYC', 'CJUS 116', 'POLI 150', 'PHIL 314', GEOG 210
    # 'MUSI 199', 'DRAM 199', 'SOCI 345', 'AGRI 295', 'ENGL 214'
    elec = ['SOCI', 'PSYC', 'CJUS 116', 'POLI 150', 'PHIL 314', \
            'GEOG 210', 'MUSI 199', 'DRAM 199', 'SOCI 345', 'AGRI 295', \
            'ENGL 214']
    for i in elec:
        if re.search(i, course):
            return True
    return False


# Check for CSCI/MATH/STAT Elective
def cms_elec(course):
    # any course that satisfies csci_elec or math_elec
    # 'STAT 380'
    if csci_elec(course):
        return True
    if math_elec(course):
        return True
    if course == 'STAT 380':
        return True
    return False


# Check for Literature Elective
def lit_elec(course):
    # 'ENGL 201', 'ENGL 202', 'ENGL 210', 'ENGL 211', 'ENGL 212', 'ENGL 213'
    # 'ENGL 214', 'ENGL 215'
    elec = ['ENGL 201', 'ENGL 202', 'ENGL 210', 'ENGL 211', 'ENGL 212', \
            'ENGL 213', 'ENGL 214', 'ENGL 215']
    if course in elec:
        return True
    else:
        return False


# Check for History Elective
def hist_elec(course):
    # 'HIST'
    if re.search('HIST', course):
        return True
    else:
        return False


# Check for Science Elective
def sci_elec(course):
    # 'BIOL 120 + LAB', 'BIOL 121 + LAB', 'CHEM 151 + CHEM 153'
    # 'PHYS 105 + LAB', 'PHYS 106 + LAB', 'PHYS 112 + LAB'
    # 'PHYS 113 + LAB'
    elec = ['BIOL 120', 'BIOL 121', 'CHEM 151', 'CHEM 153', 'PHYS 105', \
            'PHYS 106', 'PHYS 112', 'PHYS 113']
    if course in elec:
        return True
    else:
        return False


# check for matching course and lab
def sci_check(lab, course):
    # All labs have the same ID as their course coutnerpart
    if lab == course:
        return True
    # special case for 'CHEM 151', whose lab has a different ID
    elif lab == 'CHEM 153' and course == 'CHEM 151':
        return True
    else:
        return False


# Check for Math Elective
def math_elec(course):
    # 'MATH 325', 'MATH 340', 'MATH 348', 'MATH 350', 'MATH 360', 'MATH 425'
    elec = ['MATH 325', 'MATH 340', 'MATH 348', 'MATH 350', 'MATH 360', 'MATH 425']

    # -----ADD MATH 392 AS ELECTIVE, MUST HAVE MATH 260--------

    if course in elec:
        return True
    else:
        return False


# Check for possible course substitutions
def check_subs(core_course_stack, course):
    pass


# Finds the current semester
def get_curr_sem():
    month = datetime.now().month
    year = datetime.now().year

    if month >= 1:
        if month >= 5:
            if month <= 8:
                if month == 12:
                    return "Winter " + str(year)
                else:
                    return "Fall " + str(year)
            else:
                return "Summer " + str(year)
        else:
            return "Spring " + str(year)


# Finds common courses between dictionaries
def find_courses(semester, core_courses, electives):
    common_keys = list(set(semester).intersection(core_courses)), list(set(semester).intersection(electives))
    for line in common_keys:
        for key in line:
            try:
                if bool(core_courses[key][0]):
                    semester[key] = core_courses[key][0]
            except KeyError:
                if bool(electives[key][0]):
                    semester[key] = electives[key][0]


# method to add 6th index to elements missing it
def append_dict_elems(semester):
    keys = semester.keys()
    for key in semester:
        # Some elements already have 6 indices
        # check for length of course
        # if course length is 6, do not append
        if len(semester[key]) < 6:
            semester[key].append('')


# Merges index columns and styles the cells
def format_cells(wb, ws, start, end, start_cell, end_cell, col_size):
    # finds index columns and merges them
    ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
    # grabs range of cells that align with merged indices, and creates a border
    cells = list(ws[start_cell:end_cell])
    i = 0
    for row in cells:
        j = 0
        row_size = len(row) - 1
        is_complete = True
        in_progress = False
        is_open = False
        # find all courses that either have not been taken yet or are in progress
        for cell in row:
            if cell.value == None:
                if not row[j] == row[-1]:
                    is_complete = False
            if cell.value == "In progress":
                in_progress = True
            if cell.value == "Eligible":
                is_open = True
            j += 1
        j = 0
        for cell in row:
            # for all taken or in progress courses, mark green and yellow respectively
            if is_complete:
                # for courses in progress, mark yellow
                if in_progress:
                    # if first index, then we are at the top border cell
                    if i == 0:
                        # if first index, we are at the left border cell
                        if j == 0:
                            cells[i][j].border = top_left
                            cells[i][j].fill = yellowFill
                        # if last index, then we are at the right border cell
                        elif j + 1 == row_size:
                            cells[i][j].border = top_right
                            cells[i][j].fill = yellowFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = top
                            cells[i][j].fill = yellowFill
                    # if last index, then we are at the bottom border cell
                    elif i + 1 == col_size:
                        if j == 0:
                            cells[i][j].border = bottom_left
                            cells[i][j].fill = yellowFill
                        elif j + 1 == row_size:
                            cells[i][j].border = bottom_right
                            cells[i][j].fill = yellowFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = bottom
                            cells[i][j].fill = yellowFill
                    elif j == 0:
                        cells[i][j].border = left
                        cells[i][j].fill = yellowFill
                    elif j + 1 == row_size:
                        cells[i][j].border = right
                        cells[i][j].fill = yellowFill
                    # for inner cells, make border thin
                    elif not row[j] == row[-1]:
                        cells[i][j].border = other
                        cells[i][j].fill = yellowFill
                # for completed courses, mark green
                else:
                    if i == 0:
                        if j == 0:
                            cells[i][j].border = top_left
                            cells[i][j].fill = greenFill
                        elif j + 1 == row_size:
                            cells[i][j].border = top_right
                            cells[i][j].fill = greenFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = top
                            cells[i][j].fill = greenFill
                    elif i + 1 == col_size:
                        if j == 0:
                            cells[i][j].border = bottom_left
                            cells[i][j].fill = greenFill
                        elif j + 1 == row_size:
                            cells[i][j].border = bottom_right
                            cells[i][j].fill = greenFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = bottom
                            cells[i][j].fill = greenFill
                    elif j == 0:
                        cells[i][j].border = left
                        cells[i][j].fill = greenFill
                    elif j + 1 == row_size:
                        cells[i][j].border = right
                        cells[i][j].fill = greenFill
                    elif not row[j] == row[-1]:
                        cells[i][j].border = other
                        cells[i][j].fill = greenFill
            # for all courses that are not taken, mark blue if eligible, or red if ineligible
            else:
                # For courses the student is eligible for, mark blue
                if is_open:
                    if i == 0:
                        if j == 0:
                            cells[i][j].border = top_left
                            cells[i][j].fill = blueFill
                        elif j + 1 == row_size:
                            cells[i][j].border = top_right
                            cells[i][j].fill = blueFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = top
                            cells[i][j].fill = blueFill
                    elif i + 1 == col_size:
                        if j == 0:
                            cells[i][j].border = bottom_left
                            cells[i][j].fill = blueFill
                        elif j + 1 == row_size:
                            cells[i][j].border = bottom_right
                            cells[i][j].fill = blueFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = bottom
                            cells[i][j].fill = blueFill
                    elif j == 0:
                        cells[i][j].border = left
                        cells[i][j].fill = blueFill
                    elif j + 1 == row_size:
                        cells[i][j].border = right
                        cells[i][j].fill = blueFill
                    elif not row[j] == row[-1]:
                        cells[i][j].border = other
                        cells[i][j].fill = blueFill
                # For all other courses, mark red
                else:
                    if i == 0:
                        if j == 0:
                            cells[i][j].border = top_left
                            cells[i][j].fill = redFill
                        elif j + 1 == row_size:
                            cells[i][j].border = top_right
                            cells[i][j].fill = redFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = top
                            cells[i][j].fill = redFill
                    elif i + 1 == col_size:
                        if j == 0:
                            cells[i][j].border = bottom_left
                            cells[i][j].fill = redFill
                        elif j + 1 == row_size:
                            cells[i][j].border = bottom_right
                            cells[i][j].fill = redFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = bottom
                            cells[i][j].fill = redFill
                    elif j == 0:
                        cells[i][j].border = left
                        cells[i][j].fill = redFill
                    elif j + 1 == row_size:
                        cells[i][j].border = right
                        cells[i][j].fill = redFill
                    elif not row[j] == row[-1]:
                        cells[i][j].border = other
                        cells[i][j].fill = redFill

            j += 1
            if j == len(row):
                break
        i += 1


def main(courses, name):
    print("Generating Advisory Report for " + name + "...")
    # Dictionary to hold core course stack, as we traverse through course matrix,
    # we'll set dict values to true as we find them
    # MAKE CHANGE IN DICTIONARY STRUCTURE:
    # COMBINE PHIL 450 AND PHIL 275
    core_course_stack = {"MATH 260": [''], "CSCI 101": [''], "CSCI 150": [''], "CSCI 151": [''],
                         "ENGL 110": [''], "MATH 261": [''], "CSCI 250": [''], "CSCI 251": [''],
                         "ENGL 111": [''], "MATH 280": [''], "CSCI 287": [''], "CSCI 303": [''],
                         "ENGL 342": [''], "CSCI 281": [''], "CSCI 356": [''], "CSCI 296": [''],
                         "PHIL 450": [''], "PHIL 275": [''], "STAT 340": [''], "CSCI 392": [''],
                         "CSCI 487": [''], "CSCI 445": [''], "CSCI 358": [''], "CSCI 489": [''],
                         "CSCI 400": [''], "CSCI 471": [''], "CSCI 493": [''], "CSCI 485": [''],
                         "CSCI 494": [''], "GEEN 310": [''], "HPER 170": [''],
                         "MATH 284": [''], "STAT 330": ['']}
    # likewise, a dictionary to hold elective courses
    elec_stack = {"SS": [''], "HIST": [''], "LIT": [''], "SCI 1": [''], "SCI 2": [''],
                  "MATH": [''], "CMS": [''], "CSCI 1": [''], "CSCI 2": [''], "GS": [''],
                  "FREE 1": [''], "FREE 2": [''], "SCI 1 LAB": [''], "SCI 2 LAB": ['']}

    # variable to hold the current semester
    curr_sem = get_curr_sem()

    # Get credits from transcript
    core_credits = 0
    total_credits = 0

    for c in courses:
        total_credits += float(c[3])
        if c[0] in core_course_stack:
            core_credits += float(c[3])

    # holds science courses so that we can compare them
    # match a course to a lab
    sci = {}
    unused = {}

    # loop through matrix to find courses matching requirements
    for c in courses:
        passed_course = False
        course = c[0]
        cname = c[1]
        grade = c[2]
        credit = float(c[3])
        term = c[4]
        abbr = course[:5]
        # search for core class requirements first
        if course in core_course_stack:
            # for certain courses, 'D' is a passing grade
            if course == "PHIL 450" or course == "PHIL 275" or course == "ENGL 342":
                passed_course = passing_grade_alt(grade, term, curr_sem)
            # for all other courses, 'C' is the minimum
            else:
                passed_course = passing_grade(grade, term, curr_sem)
            if passed_course:
                core_course_stack[course] = [c]
        # for non-core courses
        else:
            # check global studies elective
            if gs_elec(course) and not bool(elec_stack["GS"][0]):
                elec_stack["GS"] = [c]
                elec_stack["GS"].append("")
            # check literature elective
            elif lit_elec(course) and not bool(elec_stack["LIT"][0]):
                elec_stack["LIT"] = [c]
                elec_stack["LIT"].append("")
            # check social science elective
            elif ss_elec(course) and not bool(elec_stack["SS"][0]):
                elec_stack["SS"] = [c]
                elec_stack["SS"].append("")
            # check csci electives
            elif csci_elec(course) and not bool(elec_stack["CSCI 2"][0]):
                if bool(elec_stack["CSCI 1"][0]):
                    elec_stack["CSCI 2"] = [c]
                    elec_stack["CSCI 2"].append("")
                else:
                    elec_stack["CSCI 1"] = [c]
                    elec_stack["CSCI 1"].append("")
            # check math elective
            elif math_elec(course) and not bool(elec_stack["MATH"][0]):
                elec_stack["MATH"] = [c]
                elec_stack["MATH"].append("")
            # check history elective
            elif hist_elec(course) and not bool(elec_stack["HIST"][0]):
                elec_stack["HIST"] = [c]
                elec_stack["HIST"].append("")
            # check science electives
            elif sci_elec(course):
                # check if course is lab class
                if credit == 1:
                    # search for main course
                    for itr in sci:
                        # if main course matches lab, then science is satisfied
                        if sci_check(course, itr):
                            if not bool(elec_stack["SCI 1 LAB"][0]):
                                elec_stack["SCI 1 LAB"] = [c]
                                elec_stack["SCI 1 LAB"].append("")
                            elif not bool(elec_stack["SCI 2 LAB"][0]):
                                elec_stack["SCI 2 LAB"] = [c]
                                elec_stack["SCI 2 LAB"].append("")
                # If not a lab class, then course must be main course
                else:
                    # set main course to array to compare later
                    if not course in sci:
                        sci[course] = c
                    # search for matching lab course
                    if elec_stack["SCI 1"][0] == "":
                        elec_stack["SCI 1"] = [c]
                        elec_stack["SCI 1"].append("")
                    elif elec_stack["SCI 2"][0] == "":
                        elec_stack["SCI 2"] = [c]
                        elec_stack["SCI 2"].append("")
            # csci/math/stat elective
            elif cms_elec(course) and not bool(elec_stack["CMS"][0]):
                elec_stack["CMS"] = [c]
                elec_stack["CMS"].append("")
            # unused courses
            else:
                # exclude courses with 0 credits
                if not bool(elec_stack["FREE 1"][0]) and credit > 0:
                    elec_stack["FREE 1"] = [c]
                    elec_stack["FREE 1"].append("")
                elif not bool(elec_stack["FREE 2"][0]) and credit > 0:
                    elec_stack["FREE 2"] = [c]
                    elec_stack["FREE 2"].append("")
                else:
                    unused[course] = c
                    unused[course].append("")

    # dictionaries that hold courses according to recommended semesters
    fr1 = {"MATH 260": ['MATH 260', 'Calculus I', '', '', '', ''],
           "CSCI 101": ['CSCI 101', 'Intro to Computer Sci', '', '', '', ''],
           "CSCI 150": ['CSCI 150', 'Programming I', '', '', '', ''],
           "CSCI 151": ['CSCI 151', 'Programming I Lab', '', '', '', ''],
           "ENGL 110": ['ENGL 110', 'Composition I', '', '', '', ''],
           "HPER 170": ['HPER 170', 'Health & Wellness', '', '', '', '']}
    fr2 = {"MATH 261": ['MATH 261', 'Calculus II', '', '', '', ''],
           "CSCI 250": ['CSCI 250', 'Programming II', '', '', '', ''],
           "CSCI 251": ['CSCI 251', 'Programming II Lab', '', '', '', ''],
           "ENGL 111": ['ENGL 111', 'Composition II', '', '', '', ''],
           "HIST": ['History Elective', '', '', '', ''], "SS": ['Social Science Elective', '', '', '', '']}
    so1 = {"GS": ['Global Studies Elective', '', '', '', ''],
           "MATH 280": ['MATH 280', 'Discr Math for Comp Sci', '', '', '', ''],
           "CSCI 287": ['CSCI 287', 'Data Structures', '', '', '', ''],
           "CSCI 303": ['CSCI 303', 'Computer Org & Architecture', '', '', '', ''],
           "ENGL 342": ['ENGL 342', 'Technical Communication', '', '', '', '']}
    so2 = {"LIT": ['Literature Elective', '', '', '', ''],
           "CSCI 281": ['CSCI 281', 'Discrete Structures', '', '', '', ''],
           "CSCI 356": ['CSCI 356', 'Database Systems', '', '', '', ''],
           "CSCI 296": ['CSCI 296', 'Web Programming', '', '', '', ''],
           "PHIL 450": ['PHIL 450', 'Applied Ethics', '', '', '', '']}
    jr1 = {"STAT 340": ['STAT 340', 'Prob & Stats for CS', '', '', '', ''],
           "CSCI 392": ['CSCI 392', 'Algorithms & Adv Data Structur', '', '', '', ''],
           "CSCI 487": ['CSCI 487', 'Software Design And Developmen', '', '', '', ''],
           "CSCI 445": ['CSCI 445', 'Computer Communication Network', '', '', '', ''],
           "SCI 1": ['Science Elective', '', '', '', ''], "SCI 1 LAB": ['Science Elective Lab', '', '', '', '']}
    jr2 = {"CMS": ['CSCI/MATH/STAT Elective', '', '', '', ''],
           "CSCI 358": ['CSCI 358', 'Introd Information Assurance', '', '', '', ''],
           "CSCI 489": ['CSCI 489', 'Operating Systems', '', '', '', ''],
           "CSCI 400": ['CSCI 400', 'Computer Science Seminar', '', '', '', ''],
           "FREE 1": ['Free Elective', '', '', '', '']}
    sr1 = {"CSCI 471": ['CSCI 471', 'Parallel & Distributed Program', '', '', '', ''],
           "CSCI 1": ['CSCI Elective', '', '', '', ''],
           "FREE 2": ['Free Elective', '', '', '', ''], "CSCI 493": ['CSCI 493', 'Senior Project I', '', '', '', ''],
           "SCI 2": ['Science Elective', '', '', '', ''], "SCI 2 LAB": ['Science Elective Lab', '', '', '', '']}
    sr2 = {"MATH": ['Math Elective', '', '', '', ''], "CSCI 2": ['CSCI Elective', '', '', '', ''],
           "CSCI 485": ['CSCI 485', 'Programming Languages', '', '', '', ''],
           "CSCI 494": ['CSCI 494', 'Senior Project II', '', '', '', '']}

    # find the common keys between semester dictionaries and course stacks
    common_keys = list(set(fr1).intersection(core_course_stack))
    for key in common_keys:
        if bool(core_course_stack[key][0]):
            fr1[key] = core_course_stack[key][0]

    find_courses(fr2, core_course_stack, elec_stack)
    find_courses(so1, core_course_stack, elec_stack)
    # Allow MATH 284 to replace MATH 280 if MATH 280 is missing
    if so1['MATH 280'][2] == "":
        if not core_course_stack['MATH 284'] == [""]:
            so1['MATH 280'] = [""]
            for i in core_course_stack["MATH 284"]:
                for j in i:
                    so1['MATH 280'].append(j)
            so1['MATH 280'].pop(0)
            so1['MATH 280'].append("MATH 280 Sub")
    find_courses(so2, core_course_stack, elec_stack)
    # Allow PHIL 275 as an alternative to PHIL 450 (NOT A SUB)
    if so2['PHIL 450'][2] == "":
        if not core_course_stack['PHIL 275'] == [""]:
            so2['PHIL 450'] = [""]
            for i in core_course_stack["PHIL 275"]:
                for j in i:
                    so2['PHIL 450'].append(j)
            so2['PHIL 450'].pop(0)
            so2['PHIL 450'].append("")
    find_courses(jr1, core_course_stack, elec_stack)
    # Allow STAT 330 to replace STAT 340 if STAT 340 is missing
    if jr1['STAT 340'][2] == "":
        if not core_course_stack['STAT 330'] == [""]:
            jr1["STAT 340"] = [""]
            for i in core_course_stack['STAT 330']:
                for j in i:
                    jr1['STAT 340'].append(j)
            jr1['STAT 340'].pop(0)
            jr1['STAT 340'].append("STAT 340 Sub")
    find_courses(jr2, core_course_stack, elec_stack)
    find_courses(sr1, core_course_stack, elec_stack)
    find_courses(sr2, core_course_stack, elec_stack)

    # adds 6th index for elements missing 6th index
    append_dict_elems(fr1)
    append_dict_elems(fr2)
    append_dict_elems(so1)
    append_dict_elems(so2)
    append_dict_elems(jr1)
    append_dict_elems(jr2)
    append_dict_elems(sr1)
    append_dict_elems(sr2)

    # Math courses stack
    if bool(fr1["MATH 260"][2]):
        if bool(fr2["MATH 261"][2]):
            if bool(so2["CSCI 281"][2]) and not bool(jr1["STAT 340"][2]):
                jr1['STAT 340'][5] = "Eligible"
        else:
            fr2['MATH 261'][5] = "Eligible"
    else:
        fr1['MATH 260'][5] = "Eligible"

    if not bool(so1["MATH 280"][2]):
        so1['MATH 280'][5] = "Eligible"

    # CSCI courses stack
    if not bool(fr1["CSCI 101"][2]):
        fr1['CSCI 101'][5] = "Eligible"

    if not bool(fr1["CSCI 150"][2]):
        fr1['CSCI 150'][5] = "Eligible"

    if not bool(fr1["CSCI 151"][2]):
        fr1['CSCI 151'][5] = "Eligible"

    if bool(so1["MATH 280"][2]):
        if bool(so2["CSCI 281"][2]):
            if bool(so1["CSCI 287"][2]) and not bool(jr1["CSCI 392"][2]):
                jr1['CSCI 392'][5] = "Eligible"
        else:
            so2['CSCI 281'][5] = "Eligible"

    if bool(fr1["CSCI 150"][2]) and bool(fr1["CSCI 151"][2]):
        if not bool(fr2["CSCI 250"][2]):
            fr2['CSCI 250'][5] = "Eligible"

        if not bool(fr2["CSCI 251"][2]):
            fr2['CSCI 251'][5] = "Eligible"

        if bool(fr2["CSCI 250"][2]) and bool(fr2["CSCI 251"][2]):
            if not bool(so2["CSCI 296"][2]):
                so2['CSCI 296'][5] = "Eligible"

            if not bool(so2["CSCI 356"][2]):
                so2['CSCI 356'][5] = "Eligible"

            if not bool(jr2["CSCI 358"][2]):
                jr2['CSCI 358'][5] = "Eligible"

            if bool(so1["CSCI 287"][2]):
                if not bool(jr1["CSCI 487"][2]):
                    jr1['CSCI 487'][5] = "Eligible"
                if not bool(sr1["CSCI 471"][2]):
                    sr1['CSCI 471'][5] = "Eligible"
                if not bool(sr2["CSCI 485"][2]):
                    sr2['CSCI 485'][5] = "Eligible"
            else:
                so1['CSCI 287'][5] = "Eligible"

            if bool(so1["CSCI 303"][2]):
                if not bool(jr2["CSCI 489"][2]):
                    jr2['CSCI 489'][5] = "Eligible"
                if not bool(jr1["CSCI 445"][2]):
                    jr1['CSCI 445'][5] = "Eligible"
            else:
                so1['CSCI 303'][5] = "Eligible"

    if bool(jr2["CSCI 400"][2]):
        if bool(sr1["CSCI 493"][2]):
            if not bool(sr2["CSCI 494"][2]):
                sr2['CSCI 494'][5] = "Eligible"
        else:
            sr1['CSCI 493'][5] = "Eligible"
    elif total_credits > 75:
        jr2['CSCI 400'][5] = "Eligible"

    # English courses stack
    if bool(fr1["ENGL 110"][2]):
        if bool(fr2["ENGL 111"][2]):
            if not bool(so1["ENGL 342"][2]):
                so1['ENGL 342'][5] = "Eligible"
        else:
            fr2['ENGL 111'][5] = "Eligible"
    else:
        fr1['ENGL 110'][5] = "Eligible"

    # Miscellaneous courses stack
    if not bool(so2["PHIL 450"][2]):
        so2['PHIL 450'][5] = "Eligible"
    if not bool(fr1["HPER 170"][2]):
        fr1['HPER 170'][5] = "Eligible"

    # Elective courses stack
    if not bool(fr2["SS"][2]):
        fr2['SS'][5] = "Eligible"
    if not bool(fr2["HIST"][2]):
        fr2['HIST'][5] = "Eligible"
    if not bool(so2["LIT"][2]):
        so2['LIT'][5] = "Eligible"
    if not bool(jr1["SCI 1"][2]):
        jr1['SCI 1'][5] = "Eligible"
    if not bool(sr1["SCI 2"][2]):
        sr1['SCI 2'][5] = "Eligible"
    if not bool(jr1["SCI 1 LAB"][2]):
        jr1['SCI 1 LAB'][5] = "Eligible"
    if not bool(sr1["SCI 2 LAB"][2]):
        sr1['SCI 2 LAB'][5] = "Eligible"
    if not bool(sr2["MATH"][2]):
        sr2['MATH'][5] = "Eligible"
    if not bool(jr2["CMS"][2]):
        jr2['CMS'][5] = "Eligible"
    if not bool(sr1["CSCI 1"][2]):
        sr1['CSCI 1'][5] = "Eligible"
    if not bool(sr2["CSCI 2"][2]):
        sr2['CSCI 2'][5] = "Eligible"
    if not bool(so1["GS"][2]):
        so1['GS'][5] = "Eligible"
    if not bool(jr2["FREE 1"][2]):
        jr2['FREE 1'][5] = "Eligible"
    if not bool(sr1["FREE 2"][2]):
        sr1['FREE 2'][5] = "Eligible"

    # test code for course structure, remove when finished
    print("------------------------------")
    print("Freshman: Semester 1")
    print("------------------------------")
    keys = list(fr1.keys())
    for key in keys:
        print(fr1[key])
    print("------------------------------")
    print("Freshman: Semester 2")
    print("------------------------------")
    keys = list(fr2.keys())
    for key in keys:
        print(fr2[key])
    print("------------------------------")
    print("Sophomore: Semester 1")
    print("------------------------------")
    keys = list(so1.keys())
    for key in keys:
        print(so1[key])
    print("------------------------------")
    print("Sophomore: Semester 2")
    print("------------------------------")
    keys = list(so2.keys())
    for key in keys:
        print(so2[key])
    print("------------------------------")
    print("Junior: Semester 1")
    print("------------------------------")
    keys = list(jr1.keys())
    for key in keys:
        print(jr1[key])
    print("------------------------------")
    print("Junior: Semester 2")
    print("------------------------------")
    keys = list(jr2.keys())
    for key in keys:
        print(jr2[key])
    print("------------------------------")
    print("Senior: Semester 1")
    print("------------------------------")
    keys = list(sr1.keys())
    for key in keys:
        print(sr1[key])
    print("------------------------------")
    print("Senior: Semester 2")
    print("------------------------------")
    keys = list(sr2.keys())
    for key in keys:
        print(sr2[key])
    print("------------------------------")
    print("Unused Courses")
    print("------------------------------")
    keys = list(unused.keys())
    for key in keys:
        print(unused[key])

    # create index column for excel file
    index = []
    for i in range(len(fr1)):
        index.append('Freshman: Semester 1')
    for i in range(len(fr2)):
        index.append('Freshman: Semester 2')
    for i in range(len(so1)):
        index.append('Sophomore: Semester 1')
    for i in range(len(so2)):
        index.append('Sophomore: Semester 2')
    for i in range(len(jr1)):
        index.append('Junior: Semester 1')
    for i in range(len(jr2)):
        index.append('Junior: Semester 2')
    for i in range(len(sr1)):
        index.append('Senior: Semester 1')
    for i in range(len(sr2)):
        index.append('Senior: Semester 2')
    for i in range(len(unused)):
        index.append('Unused Courses')

    # restructure courses to group by suggested semester for excel file
    courses.clear()
    fr1 = list(fr1.values())
    for course in fr1:
        courses.append(course)
    fr2 = list(fr2.values())
    for course in fr2:
        courses.append(course)
    so1 = list(so1.values())
    for course in so1:
        courses.append(course)
    so2 = list(so2.values())
    for course in so2:
        courses.append(course)
    jr1 = list(jr1.values())
    for course in jr1:
        courses.append(course)
    jr2 = list(jr2.values())
    for course in jr2:
        courses.append(course)
    sr1 = list(sr1.values())
    for course in sr1:
        courses.append(course)
    sr2 = list(sr2.values())
    for course in sr2:
        courses.append(course)
    unused = list(unused.values())
    for course in unused:
        courses.append(course)

    # structure index and coursesd into a dataframe
    df = pd.DataFrame(courses,
                      index,
                      columns=['ID', 'Name', 'Grade', 'Credits', 'Semester', 'Notes'])
    # print to excel file
    path = 'students\\' + name + '\course_log.xlsx'
    df.to_excel(path, sheet_name='Transcript')
    # modify created excel file
    wb = load_workbook(filename=path)
    ws = wb.active

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 7
    ws.column_dimensions['F'].width = 12

    for row in ws:
        cells = [row[0], row[3], row[5]]
        for cell in cells:
            if cell == cells[0]:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center')

    col_size = len(courses)

    start = 2
    end = len(fr1) + 1

    format_cells(wb, ws, start, end, 'B2', 'G7', col_size)

    start = end + 1
    end += len(fr2)

    format_cells(wb, ws, start, end, 'B8', 'G13', col_size)

    start = end + 1
    end += len(so1)

    format_cells(wb, ws, start, end, 'B14', 'G18', col_size)

    start = end + 1
    end += len(so2)

    format_cells(wb, ws, start, end, 'B19', 'G23', col_size)

    start = end + 1
    end += len(jr1)

    format_cells(wb, ws, start, end, 'B24', 'G29', col_size)

    start = end + 1
    end += len(jr2)

    format_cells(wb, ws, start, end, 'B30', 'G34', col_size)

    start = end + 1
    end += len(sr1)

    format_cells(wb, ws, start, end, 'B35', 'G40', col_size)

    start = end + 1
    end += len(sr2)

    format_cells(wb, ws, start, end, 'B41', 'G44', col_size)

    start = end + 1
    end += len(unused)

    end_cell = 'F' + str(end)

    ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
    cells = list(ws['B45':end_cell])
    i = 0
    for row in cells:
        j = 0
        row_size = len(row)
        # find all courses that have not been taken yet
        for cell in row:
            # if first index, then we are at the top border cell
            if i == 0:
                # if first index, we are at the left border cell
                if j == 0:
                    cells[i][j].border = top_left
                # if last index, then we are at the right border cell
                elif j + 1 == row_size:
                    cells[i][j].border = top_right
                else:
                    cells[i][j].border = top
            # if last index, then we are at the bottom border cell
            elif i + 1 == col_size:
                if j == 0:
                    cells[i][j].border = bottom_left
                elif j + 1 == row_size:
                    cells[i][j].border = bottom_right
                else:
                    cells[i][j].border = bottom
            elif j == 0:
                cells[i][j].border = left
            elif j + 1 == row_size:
                cells[i][j].border = right
            # for inner cells, make border thin
            else:
                cells[i][j].border = other
            j += 1
        i += 1

    # save changes to excel file
    wb.save(path)
top_left = Border(left = Side(style = 'thick'),
              right = Side(style = 'thin'),
              top = Side(style = 'thick'),
              bottom = Side(style = 'thin'))
top_right = Border(left = Side(style = 'thin'),
              right = Side(style = 'thick'),
              top = Side(style = 'thick'),
              bottom = Side(style = 'thin'))
right = Border(left = Side(style = 'thin'),
              right = Side(style = 'thick'),
              top = Side(style = 'thin'),
              bottom = Side(style = 'thin'))
bottom_right = Border(left = Side(style = 'thin'),
              right = Side(style = 'thick'),
              top = Side(style = 'thin'),
              bottom = Side(style = 'thick'))
bottom = Border(left = Side(style = 'thin'),
              right = Side(style = 'thin'),
              top = Side(style = 'thin'),
              bottom = Side(style = 'thick'))
bottom_left = Border(left = Side(style = 'thick'),
              right = Side(style = 'thin'),
              top = Side(style = 'thin'),
              bottom = Side(style = 'thick'))
left = Border(left = Side(style = 'thick'),
              right = Side(style = 'thin'),
              top = Side(style = 'thin'),
              bottom = Side(style = 'thin'))
other = Border(left = Side(style = 'thin'),
              right = Side(style = 'thin'),
              top = Side(style = 'thin'),
              bottom = Side(style = 'thin'))
#global variables for cell background colors
redFill = PatternFill('solid', fgColor = 'FF0000')
greenFill = PatternFill('solid', fgColor = '00FF00')
yellowFill = PatternFill('solid', fgColor = 'FFFF00')
blueFill = PatternFill('solid', fgColor = '00BFFF')

#at least 8 csci courses, 60 credits, and csci 287 before 400-level csci courses
#method to check if a course has a passing grade where a passing grade is 'C'
def passing_grade(grade, sem, curr_sem):
    if grade <= "C" or grade == "S" or grade == "TR" or grade == "In progress" \
       or grade == "SP":
        #add a check for future courses, which should be omitted
        #from course recommendations
        #If course is from this semester, include it
        if grade == "In progress":
            if sem == curr_sem:
                return True
            #If course is from future semester, omit it
            else:
                return False
        #If course is not from this semester, then it has already
        #been taken
        else:
            return True
    #If we reached this, then we reached a failing course somehow
    else:
        return False

#method to check if a course has a passing grade where a passing grade is 'D'
def passing_grade_alt(grade, sem, curr_sem):
    if grade <= "D" or grade == "S" or grade == "TR" or grade == "In progress" \
       or grade == "SP":
        if grade == "In progress":
            if sem == curr_sem:
                return True
            else:
                return False
        else:
            return True
    else:
        return False

#Check for CSCI Elective
def csci_elec(course):
    #'CSCI 312', 'CSCI 361', 'CSCI 389', 'CSCI 396', 'CSCI 402', 'CSCI 450'
    #'CSCI 451', 'CSCI 453', 'CSCI 495', 'CSCI 496'
    elec = ['CSCI 312', 'CSCI 361', 'CSCI 389', 'CSCI 396', 'CSCI 402', \
            'CSCI 450', 'CSCI 451', 'CSCI 453', 'CSCI 495', 'CSCI 496']
    if course in elec:
        return True
    else:
        return False
    
#Check for Global Studies Elective
def gs_elec(course):
    #'SPAN', 'FREN', 'GERM', 'HIST 114', 'HIST 115', 'ARTS 199'
    elec = ['SPAN', 'FREN', 'GERM', 'HIST 114', 'HIST 115', 'ARTS 199']

    for i in elec:
        if re.search(i, course):
            return True
    return False

#Check for Social Sciences Elective
def ss_elec(course):
    #'SOCI', 'PSYC', 'CJUS 116', 'POLI 150', 'PHIL 314', GEOG 210
    #'MUSI 199', 'DRAM 199', 'SOCI 345', 'AGRI 295', 'ENGL 214'
    elec = ['SOCI', 'PSYC', 'CJUS 116', 'POLI 150', 'PHIL 314', \
            'GEOG 210', 'MUSI 199', 'DRAM 199', 'SOCI 345', 'AGRI 295', \
            'ENGL 214']
    for i in elec:
        if re.search(i, course):
            return True
    return False

#Check for CSCI/MATH/STAT Elective
def cms_elec(course):
    #any course that satisfies csci_elec or math_elec
    #'STAT 380'
    if csci_elec(course):
        return True
    if math_elec(course):
        return True
    if course == 'STAT 380':
        return True
    return False

#Check for Literature Elective
def lit_elec(course):
    #'ENGL 201', 'ENGL 202', 'ENGL 210', 'ENGL 211', 'ENGL 212', 'ENGL 213'
    #'ENGL 214', 'ENGL 215'
    elec = ['ENGL 201', 'ENGL 202', 'ENGL 210', 'ENGL 211', 'ENGL 212', \
            'ENGL 213', 'ENGL 214', 'ENGL 215']
    if course in elec:
        return True
    else:
        return False
    
#Check for History Elective
def hist_elec(course):
    #'HIST'
    if re.search('HIST', course):
        return True
    else:
        return False
    
#Check for Science Elective
def sci_elec(course):
    #'BIOL 120 + LAB', 'BIOL 121 + LAB', 'CHEM 151 + CHEM 153'
    #'PHYS 105 + LAB', 'PHYS 106 + LAB', 'PHYS 112 + LAB'
    #'PHYS 113 + LAB'
    elec = ['BIOL 120', 'BIOL 121', 'CHEM 151', 'CHEM 153', 'PHYS 105', \
            'PHYS 106', 'PHYS 112', 'PHYS 113']
    if course in elec:
        return True
    else:
        return False
    
#check for matching course and lab
def sci_check(lab, course):
    #All labs have the same ID as their course coutnerpart
    if lab == course:
        return True
    #special case for 'CHEM 151', whose lab has a different ID
    elif lab == 'CHEM 153' and course == 'CHEM 151':
        return True
    else:
        return False
    
#Check for Math Elective
def math_elec(course):
    #'MATH 325', 'MATH 340', 'MATH 348', 'MATH 350', 'MATH 360', 'MATH 425'
    elec = ['MATH 325', 'MATH 340', 'MATH 348', 'MATH 350', 'MATH 360', 'MATH 425']

    #-----ADD MATH 392 AS ELECTIVE, MUST HAVE MATH 260--------

    if course in elec:
        return True
    else:
        return False
    
#Check for possible course substitutions
def check_subs(core_course_stack, course):
    pass
    
#Finds the current semester
def get_curr_sem():
    month = datetime.now().month
    year = datetime.now().year

    if month >= 1:
        if month >= 5:
            if month <= 8:
                if month == 12:
                    return "Winter " + str(year)
                else:
                    return "Fall " + str(year)
            else:
                return "Summer " + str(year)
        else:
            return "Spring " + str(year)

#Finds common courses between dictionaries
def find_courses(semester, core_courses, electives):
    common_keys = list(set(semester).intersection(core_courses)), list(set(semester).intersection(electives))
    for line in common_keys:
        for key in line:
            try:
                if bool(core_courses[key][0]):
                    semester[key] = core_courses[key][0]
            except KeyError:
                if bool(electives[key][0]):
                    semester[key] = electives[key][0]

#method to add 6th index to elements missing it
def append_dict_elems(semester):
    keys = semester.keys()
    for key in semester:
        #Some elements already have 6 indices
        #check for length of course
        #if course length is 6, do not append
        if len(semester[key]) < 6:
            semester[key].append('')

#Merges index columns and styles the cells
def format_cells(wb, ws, start, end, start_cell, end_cell, col_size):
    #finds index columns and merges them
    ws.merge_cells(start_row = start, start_column = 1, end_row = end, end_column = 1)
    #grabs range of cells that align with merged indices, and creates a border
    cells = list(ws[start_cell:end_cell])
    i = 0
    for row in cells:
        j = 0
        row_size = len(row) - 1
        is_complete = True
        in_progress = False
        is_open = False
        #find all courses that either have not been taken yet or are in progress
        for cell in row:
            if cell.value == None:
                if not row[j] == row[-1]:
                    is_complete = False
            if cell.value == "In progress":
                in_progress = True
            if cell.value == "Eligible":
                is_open = True
            j += 1
        j = 0
        for cell in row:
            #for all taken or in progress courses, mark green and yellow respectively
            if is_complete:
                #for courses in progress, mark yellow
                if in_progress:
                    #if first index, then we are at the top border cell
                    if i == 0:
                        #if first index, we are at the left border cell
                        if j == 0:
                            cells[i][j].border = top_left
                            cells[i][j].fill = yellowFill
                        #if last index, then we are at the right border cell
                        elif j+1 == row_size:
                            cells[i][j].border = top_right
                            cells[i][j].fill = yellowFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = top
                            cells[i][j].fill = yellowFill
                    #if last index, then we are at the bottom border cell
                    elif i+1 == col_size:
                        if j == 0:
                            cells[i][j].border = bottom_left
                            cells[i][j].fill = yellowFill
                        elif j+1 == row_size:
                            cells[i][j].border = bottom_right
                            cells[i][j].fill = yellowFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = bottom
                            cells[i][j].fill = yellowFill
                    elif j == 0:
                        cells[i][j].border = left
                        cells[i][j].fill = yellowFill
                    elif j+1 == row_size:
                        cells[i][j].border = right
                        cells[i][j].fill = yellowFill
                    #for inner cells, make border thin
                    elif not row[j] == row[-1]:
                        cells[i][j].border = other
                        cells[i][j].fill = yellowFill
                #for completed courses, mark green
                else:
                    if i == 0:
                        if j == 0:
                            cells[i][j].border = top_left
                            cells[i][j].fill = greenFill
                        elif j+1 == row_size:
                            cells[i][j].border = top_right
                            cells[i][j].fill = greenFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = top
                            cells[i][j].fill = greenFill
                    elif i+1 == col_size:
                        if j == 0:
                            cells[i][j].border = bottom_left
                            cells[i][j].fill = greenFill
                        elif j+1 == row_size:
                            cells[i][j].border = bottom_right
                            cells[i][j].fill = greenFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = bottom
                            cells[i][j].fill = greenFill
                    elif j == 0:
                        cells[i][j].border = left
                        cells[i][j].fill = greenFill
                    elif j+1 == row_size:
                        cells[i][j].border = right
                        cells[i][j].fill = greenFill
                    elif not row[j] == row[-1]:
                        cells[i][j].border = other
                        cells[i][j].fill = greenFill
            #for all courses that are not taken, mark blue if eligible, or red if ineligible
            else:
                #For courses the student is eligible for, mark blue
                if is_open:
                    if i == 0:
                        if j == 0:
                            cells[i][j].border = top_left
                            cells[i][j].fill = blueFill
                        elif j+1 == row_size:
                            cells[i][j].border = top_right
                            cells[i][j].fill = blueFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = top
                            cells[i][j].fill = blueFill
                    elif i+1 == col_size:
                        if j == 0:
                            cells[i][j].border = bottom_left
                            cells[i][j].fill = blueFill
                        elif j+1 == row_size:
                            cells[i][j].border = bottom_right
                            cells[i][j].fill = blueFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = bottom
                            cells[i][j].fill = blueFill
                    elif j == 0:
                        cells[i][j].border = left
                        cells[i][j].fill = blueFill
                    elif j+1 == row_size:
                        cells[i][j].border = right
                        cells[i][j].fill = blueFill
                    elif not row[j] == row[-1]:
                        cells[i][j].border = other
                        cells[i][j].fill = blueFill
                #For all other courses, mark red
                else:
                    if i == 0:
                        if j == 0:
                            cells[i][j].border = top_left
                            cells[i][j].fill = redFill
                        elif j+1 == row_size:
                            cells[i][j].border = top_right
                            cells[i][j].fill = redFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = top
                            cells[i][j].fill = redFill
                    elif i+1 == col_size:
                        if j == 0:
                            cells[i][j].border = bottom_left
                            cells[i][j].fill = redFill
                        elif j+1 == row_size:
                            cells[i][j].border = bottom_right
                            cells[i][j].fill = redFill
                        elif not row[j] == row[-1]:
                            cells[i][j].border = bottom
                            cells[i][j].fill = redFill
                    elif j == 0:
                        cells[i][j].border = left
                        cells[i][j].fill = redFill
                    elif j+1 == row_size:
                        cells[i][j].border = right
                        cells[i][j].fill = redFill
                    elif not row[j] == row[-1]:
                        cells[i][j].border = other
                        cells[i][j].fill = redFill

            j += 1
            if j == len(row):
                break
        i += 1

def main(courses, name):
    print("Generating Advisory Report for " + name + "...")
    #Dictionary to hold core course stack, as we traverse through course matrix,
    #we'll set dict values to true as we find them
    #MAKE CHANGE IN DICTIONARY STRUCTURE:
        #COMBINE PHIL 450 AND PHIL 275
    core_course_stack = {"MATH 260": [''], "CSCI 101": [''], "CSCI 150": [''], "CSCI 151": [''],
                         "ENGL 110": [''], "MATH 261": [''], "CSCI 250": [''], "CSCI 251": [''],
                         "ENGL 111": [''], "MATH 280": [''], "CSCI 287": [''], "CSCI 303": [''],
                         "ENGL 342": [''], "CSCI 281": [''], "CSCI 356": [''], "CSCI 296": [''],
                         "PHIL 450": [''], "PHIL 275": [''], "STAT 340": [''], "CSCI 392": [''],
                         "CSCI 487": [''], "CSCI 445": [''], "CSCI 358": [''], "CSCI 489": [''],
                         "CSCI 400": [''], "CSCI 471": [''], "CSCI 493": [''], "CSCI 485": [''],
                         "CSCI 494": [''], "GEEN 310": [''], "HPER 170": [''],
                         "MATH 284": [''], "STAT 330": ['']}
    #likewise, a dictionary to hold elective courses
    elec_stack = {"SS": [''], "HIST": [''], "LIT": [''], "SCI 1": [''], "SCI 2": [''],
                  "MATH": [''], "CMS": [''], "CSCI 1": [''], "CSCI 2": [''], "GS": [''],
                  "FREE 1": [''], "FREE 2": [''], "SCI 1 LAB": [''], "SCI 2 LAB": ['']}

    #variable to hold the current semester
    curr_sem = get_curr_sem()

    #Get credits from transcript
    core_credits = 0
    total_credits = 0

    for c in courses:
        total_credits += float(c[3])
        if c[0] in core_course_stack:
            core_credits += float(c[3])
    
    #holds science courses so that we can compare them
    #match a course to a lab
    sci = {}
    unused = {}

    #loop through matrix to find courses matching requirements
    for c in courses:
        passed_course = False
        course = c[0]
        cname = c[1]
        grade = c[2]
        credit = float(c[3])
        term = c[4]
        abbr = course[:5]
        #search for core class requirements first
        if course in core_course_stack:
            #for certain courses, 'D' is a passing grade
            if course == "PHIL 450" or course == "PHIL 275" or course == "ENGL 342":
                passed_course = passing_grade_alt(grade, term, curr_sem)
            #for all other courses, 'C' is the minimum
            else:
                passed_course = passing_grade(grade, term, curr_sem)
            if passed_course:
                core_course_stack[course] = [c]
        #for non-core courses
        else:
            #check global studies elective
            if gs_elec(course) and not bool(elec_stack["GS"][0]):
                elec_stack["GS"] = [c]
                elec_stack["GS"].append("")
            #check literature elective
            elif lit_elec(course) and not bool(elec_stack["LIT"][0]):
                elec_stack["LIT"] = [c]
                elec_stack["LIT"].append("")
            #check social science elective
            elif ss_elec(course) and not bool(elec_stack["SS"][0]):
                elec_stack["SS"] = [c]
                elec_stack["SS"].append("")
            #check csci electives
            elif csci_elec(course) and not bool(elec_stack["CSCI 2"][0]):
                if bool(elec_stack["CSCI 1"][0]):
                    elec_stack["CSCI 2"] = [c]
                    elec_stack["CSCI 2"].append("")
                else:
                    elec_stack["CSCI 1"] = [c]
                    elec_stack["CSCI 1"].append("")
            #check math elective
            elif math_elec(course) and not bool(elec_stack["MATH"][0]):
                elec_stack["MATH"] = [c]
                elec_stack["MATH"].append("")
            #check history elective
            elif hist_elec(course) and not bool(elec_stack["HIST"][0]):
                elec_stack["HIST"] = [c]
                elec_stack["HIST"].append("")
            #check science electives
            elif sci_elec(course):
                #check if course is lab class
                if credit == 1:
                    #search for main course
                    for itr in sci:
                        #if main course matches lab, then science is satisfied
                        if sci_check(course, itr):
                            if not bool(elec_stack["SCI 1 LAB"][0]):
                                elec_stack["SCI 1 LAB"] = [c]
                                elec_stack["SCI 1 LAB"].append("")
                            elif not bool(elec_stack["SCI 2 LAB"][0]):
                                elec_stack["SCI 2 LAB"] = [c]
                                elec_stack["SCI 2 LAB"].append("")
                #If not a lab class, then course must be main course
                else:
                    #set main course to array to compare later
                    if not course in sci:
                        sci[course] = c
                    #search for matching lab course
                    if elec_stack["SCI 1"][0] == "":
                        elec_stack["SCI 1"] = [c]
                        elec_stack["SCI 1"].append("")
                    elif elec_stack["SCI 2"][0] == "":
                        elec_stack["SCI 2"] = [c]
                        elec_stack["SCI 2"].append("")
            #csci/math/stat elective
            elif cms_elec(course) and not bool(elec_stack["CMS"][0]):
                elec_stack["CMS"] = [c]
                elec_stack["CMS"].append("")
            #unused courses
            else:
                #exclude courses with 0 credits
                if not bool(elec_stack["FREE 1"][0]) and credit > 0:
                    elec_stack["FREE 1"] = [c]
                    elec_stack["FREE 1"].append("")
                elif not bool(elec_stack["FREE 2"][0]) and credit > 0:
                    elec_stack["FREE 2"] = [c]
                    elec_stack["FREE 2"].append("")
                else:
                    unused[course] = c
                    unused[course].append("")

    #dictionaries that hold courses according to recommended semesters
    fr1 = {"MATH 260": ['MATH 260', 'Calculus I', '', '', '', ''], "CSCI 101": ['CSCI 101', 'Intro to Computer Sci', '', '', '', ''],
           "CSCI 150": ['CSCI 150', 'Programming I', '', '', '', ''], "CSCI 151": ['CSCI 151', 'Programming I Lab', '', '', '', ''],
           "ENGL 110": ['ENGL 110', 'Composition I', '', '', '', ''], "HPER 170": ['HPER 170', 'Health & Wellness', '', '', '', '']}
    fr2 = {"MATH 261": ['MATH 261', 'Calculus II', '', '', '', ''], "CSCI 250": ['CSCI 250', 'Programming II', '', '', '', ''],
           "CSCI 251": ['CSCI 251', 'Programming II Lab', '', '', '', ''], "ENGL 111": ['ENGL 111', 'Composition II', '', '', '', ''],
           "HIST": ['History Elective', '', '', '', ''], "SS": ['Social Science Elective', '', '', '' ,'']}
    so1 = {"GS": ['Global Studies Elective', '', '', '', ''], "MATH 280": ['MATH 280', 'Discr Math for Comp Sci', '', '', '', ''],
           "CSCI 287": ['CSCI 287', 'Data Structures', '', '', '', ''], "CSCI 303": ['CSCI 303', 'Computer Org & Architecture', '', '', '', ''],
           "ENGL 342": ['ENGL 342', 'Technical Communication', '', '', '', '']}
    so2 = {"LIT": ['Literature Elective', '', '', '', ''], "CSCI 281": ['CSCI 281', 'Discrete Structures', '', '', '', ''],
           "CSCI 356": ['CSCI 356', 'Database Systems', '', '', '', ''], "CSCI 296": ['CSCI 296', 'Web Programming', '', '', '', ''],
           "PHIL 450": ['PHIL 450', 'Applied Ethics', '', '', '', '']}
    jr1 = {"STAT 340": ['STAT 340', 'Prob & Stats for CS', '', '', '', ''], "CSCI 392": ['CSCI 392', 'Algorithms & Adv Data Structur', '', '', '', ''],
           "CSCI 487": ['CSCI 487', 'Software Design And Developmen', '', '', '', ''], "CSCI 445": ['CSCI 445', 'Computer Communication Network', '', '', '', ''],
           "SCI 1": ['Science Elective', '', '', '', ''], "SCI 1 LAB": ['Science Elective Lab', '', '', '', '']}
    jr2 = {"CMS": ['CSCI/MATH/STAT Elective', '', '', '', ''], "CSCI 358": ['CSCI 358', 'Introd Information Assurance', '', '', '', ''],
           "CSCI 489": ['CSCI 489', 'Operating Systems', '', '', '', ''], "CSCI 400": ['CSCI 400', 'Computer Science Seminar', '', '', '', ''],
           "FREE 1": ['Free Elective', '', '', '', '']}
    sr1 = {"CSCI 471": ['CSCI 471', 'Parallel & Distributed Program', '', '', '', ''], "CSCI 1": ['CSCI Elective', '', '', '', ''],
           "FREE 2": ['Free Elective', '', '', '', ''], "CSCI 493": ['CSCI 493', 'Senior Project I', '', '', '', ''],
           "SCI 2": ['Science Elective', '', '', '', ''], "SCI 2 LAB": ['Science Elective Lab', '', '', '', '']}
    sr2 = {"MATH": ['Math Elective', '', '', '', ''], "CSCI 2": ['CSCI Elective', '', '', '', ''],
           "CSCI 485": ['CSCI 485', 'Programming Languages', '', '', '', ''], "CSCI 494": ['CSCI 494', 'Senior Project II', '', '', '', '']}

    #find the common keys between semester dictionaries and course stacks
    common_keys = list(set(fr1).intersection(core_course_stack))
    for key in common_keys:
        if bool(core_course_stack[key][0]):
            fr1[key] = core_course_stack[key][0]

    find_courses(fr2, core_course_stack, elec_stack)
    find_courses(so1, core_course_stack, elec_stack)
    #Allow MATH 284 to replace MATH 280 if MATH 280 is missing
    if so1['MATH 280'][2] == "":
        if not core_course_stack['MATH 284'] == [""]:
            so1['MATH 280'] = [""]
            for i in core_course_stack["MATH 284"]:
                for j in i:
                    so1['MATH 280'].append(j)
            so1['MATH 280'].pop(0)
            so1['MATH 280'].append("MATH 280 Sub")
    find_courses(so2, core_course_stack, elec_stack)
    #Allow PHIL 275 as an alternative to PHIL 450 (NOT A SUB)
    if so2['PHIL 450'][2] == "":
        if not core_course_stack['PHIL 275'] == [""]:
            so2['PHIL 450'] = [""]
            for i in core_course_stack["PHIL 275"]:
                for j in i:
                    so2['PHIL 450'].append(j)
            so2['PHIL 450'].pop(0)
            so2['PHIL 450'].append("")
    find_courses(jr1, core_course_stack, elec_stack)
    #Allow STAT 330 to replace STAT 340 if STAT 340 is missing
    if jr1['STAT 340'][2] == "":
        if not core_course_stack['STAT 330'] == [""]:
            jr1["STAT 340"] = [""]
            for i in core_course_stack['STAT 330']:
                for j in i:
                    jr1['STAT 340'].append(j)
            jr1['STAT 340'].pop(0)
            jr1['STAT 340'].append("STAT 340 Sub")
    find_courses(jr2, core_course_stack, elec_stack)
    find_courses(sr1, core_course_stack, elec_stack)
    find_courses(sr2, core_course_stack, elec_stack)

    #adds 6th index for elements missing 6th index
    append_dict_elems(fr1)
    append_dict_elems(fr2)
    append_dict_elems(so1)
    append_dict_elems(so2)
    append_dict_elems(jr1)
    append_dict_elems(jr2)
    append_dict_elems(sr1)
    append_dict_elems(sr2)

    #Math courses stack
    if bool(fr1["MATH 260"][2]):
        if bool(fr2["MATH 261"][2]):
            if bool(so2["CSCI 281"][2]) and not bool(jr1["STAT 340"][2]):
                jr1['STAT 340'][5] = "Eligible"
        else:       
            fr2['MATH 261'][5] = "Eligible"
    else:
        fr1['MATH 260'][5] = "Eligible"

    if not bool(so1["MATH 280"][2]):
        so1['MATH 280'][5] = "Eligible"

    #CSCI courses stack
    if not bool(fr1["CSCI 101"][2]):
        fr1['CSCI 101'][5] = "Eligible"

    if not bool(fr1["CSCI 150"][2]):
        fr1['CSCI 150'][5] = "Eligible"

    if not bool(fr1["CSCI 151"][2]):
        fr1['CSCI 151'][5] = "Eligible"

    if bool(so1["MATH 280"][2]):
        if bool(so2["CSCI 281"][2]):
            if bool(so1["CSCI 287"][2]) and not bool(jr1["CSCI 392"][2]):
                jr1['CSCI 392'][5] = "Eligible"
        else:
            so2['CSCI 281'][5] = "Eligible"

    if bool(fr1["CSCI 150"][2]) and bool(fr1["CSCI 151"][2]):
        if not bool(fr2["CSCI 250"][2]):
            fr2['CSCI 250'][5] = "Eligible"

        if not bool(fr2["CSCI 251"][2]):
            fr2['CSCI 251'][5] = "Eligible"

        if bool(fr2["CSCI 250"][2]) and bool(fr2["CSCI 251"][2]):
            if not bool(so2["CSCI 296"][2]):
               so2['CSCI 296'][5] = "Eligible"

            if not bool(so2["CSCI 356"][2]):
                so2['CSCI 356'][5] = "Eligible"

            if not bool(jr2["CSCI 358"][2]):
                jr2['CSCI 358'][5] = "Eligible"

            if bool(so1["CSCI 287"][2]):
                if not bool(jr1["CSCI 487"][2]):
                    jr1['CSCI 487'][5] = "Eligible"
                if not bool(sr1["CSCI 471"][2]):
                    sr1['CSCI 471'][5] = "Eligible"
                if not bool(sr2["CSCI 485"][2]):
                    sr2['CSCI 485'][5] = "Eligible"
            else:
                so1['CSCI 287'][5] = "Eligible"

            if bool(so1["CSCI 303"][2]):
                if not bool(jr2["CSCI 489"][2]):
                    jr2['CSCI 489'][5] = "Eligible"
                if not bool(jr1["CSCI 445"][2]):
                    jr1['CSCI 445'][5] = "Eligible"
            else:
                so1['CSCI 303'][5] = "Eligible"

    if bool(jr2["CSCI 400"][2]):
        if bool(sr1["CSCI 493"][2]):
            if not bool(sr2["CSCI 494"][2]):
                sr2['CSCI 494'][5] = "Eligible"
        else:
            sr1['CSCI 493'][5] = "Eligible"
    elif total_credits > 75:
        jr2['CSCI 400'][5] = "Eligible"

    #English courses stack
    if bool(fr1["ENGL 110"][2]):
        if bool(fr2["ENGL 111"][2]):
            if not bool(so1["ENGL 342"][2]):
                so1['ENGL 342'][5] = "Eligible"
        else:
            fr2['ENGL 111'][5] = "Eligible"
    else:
        fr1['ENGL 110'][5] = "Eligible"

    #Miscellaneous courses stack
    if not bool(so2["PHIL 450"][2]):
        so2['PHIL 450'][5] = "Eligible"
    if not bool(fr1["HPER 170"][2]):
        fr1['HPER 170'][5] = "Eligible"

    #Elective courses stack
    if not bool(fr2["SS"][2]):
        fr2['SS'][5] = "Eligible"
    if not bool(fr2["HIST"][2]):
        fr2['HIST'][5] = "Eligible"
    if not bool(so2["LIT"][2]):
        so2['LIT'][5] = "Eligible"
    if not bool(jr1["SCI 1"][2]):
        jr1['SCI 1'][5] = "Eligible"
    if not bool(sr1["SCI 2"][2]):
        sr1['SCI 2'][5] = "Eligible"
    if not bool(jr1["SCI 1 LAB"][2]):
        jr1['SCI 1 LAB'][5] = "Eligible"
    if not bool(sr1["SCI 2 LAB"][2]):
        sr1['SCI 2 LAB'][5] = "Eligible"
    if not bool(sr2["MATH"][2]):
        sr2['MATH'][5] = "Eligible"
    if not bool(jr2["CMS"][2]):
        jr2['CMS'][5] = "Eligible"
    if not bool(sr1["CSCI 1"][2]):
        sr1['CSCI 1'][5] = "Eligible"
    if not bool(sr2["CSCI 2"][2]):
        sr2['CSCI 2'][5] = "Eligible"
    if not bool(so1["GS"][2]):
        so1['GS'][5] = "Eligible"
    if not bool(jr2["FREE 1"][2]):
        jr2['FREE 1'][5] = "Eligible"
    if not bool(sr1["FREE 2"][2]):
        sr1['FREE 2'][5] = "Eligible"

    #test code for course structure, remove when finished
    print("------------------------------")
    print("Freshman: Semester 1")
    print("------------------------------")
    keys = list(fr1.keys())
    for key in keys:
        print(fr1[key])
    print("------------------------------")
    print("Freshman: Semester 2")
    print("------------------------------")
    keys = list(fr2.keys())
    for key in keys:
        print(fr2[key])
    print("------------------------------")
    print("Sophomore: Semester 1")
    print("------------------------------")
    keys = list(so1.keys())
    for key in keys:
        print(so1[key])
    print("------------------------------")
    print("Sophomore: Semester 2")
    print("------------------------------")
    keys = list(so2.keys())
    for key in keys:
        print(so2[key])
    print("------------------------------")
    print("Junior: Semester 1")
    print("------------------------------")
    keys = list(jr1.keys())
    for key in keys:
        print(jr1[key])
    print("------------------------------")
    print("Junior: Semester 2")
    print("------------------------------")
    keys = list(jr2.keys())
    for key in keys:
        print(jr2[key])
    print("------------------------------")
    print("Senior: Semester 1")
    print("------------------------------")
    keys = list(sr1.keys())
    for key in keys:
        print(sr1[key])
    print("------------------------------")
    print("Senior: Semester 2")
    print("------------------------------")
    keys = list(sr2.keys())
    for key in keys:
        print(sr2[key])
    print("------------------------------")    
    print("Unused Courses")
    print("------------------------------")
    keys = list(unused.keys())
    for key in keys:
        print(unused[key])

    #create index column for excel file
    index = []
    for i in range(len(fr1)):
        index.append('Freshman: Semester 1')
    for i in range(len(fr2)):
        index.append('Freshman: Semester 2')
    for i in range(len(so1)):
        index.append('Sophomore: Semester 1')
    for i in range(len(so2)):
        index.append('Sophomore: Semester 2')
    for i in range(len(jr1)):
        index.append('Junior: Semester 1')
    for i in range(len(jr2)):
        index.append('Junior: Semester 2')
    for i in range(len(sr1)):
        index.append('Senior: Semester 1')
    for i in range(len(sr2)):
        index.append('Senior: Semester 2')
    for i in range(len(unused)):
        index.append('Unused Courses')

    #restructure courses to group by suggested semester for excel file
    courses.clear()
    fr1 = list(fr1.values())
    for course in fr1:
        courses.append(course)
    fr2 = list(fr2.values())
    for course in fr2:
        courses.append(course)
    so1 = list(so1.values())
    for course in so1:
        courses.append(course)
    so2 = list(so2.values())
    for course in so2:
        courses.append(course)
    jr1 = list(jr1.values())
    for course in jr1:
        courses.append(course)
    jr2 = list(jr2.values())
    for course in jr2:
        courses.append(course)
    sr1 = list(sr1.values())
    for course in sr1:
        courses.append(course)
    sr2 = list(sr2.values())
    for course in sr2:
        courses.append(course)
    unused = list(unused.values())
    for course in unused:
        courses.append(course)

    #structure index and coursesd into a dataframe
    df = pd.DataFrame(courses,
                      index,
                      columns = ['ID', 'Name', 'Grade', 'Credits', 'Semester', 'Notes'])
    #print to excel file
    path = 'students\\' + name + '\course_log.xlsx'
    df.to_excel(path, sheet_name = 'Transcript')
    #modify created excel file
    wb = load_workbook(filename = path)
    ws = wb.active

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 7
    ws.column_dimensions['F'].width = 12

    for row in ws:
        cells = [row[0], row[3], row[5]]
        for cell in cells:
            if cell == cells[0]:
                cell.alignment = Alignment(wrap_text = True, horizontal = 'center', vertical = 'center')
            else:
                cell.alignment = Alignment(horizontal = 'center')

    col_size = len(courses)

    start = 2
    end = len(fr1) + 1

    format_cells(wb, ws, start, end, 'B2', 'G7', col_size)
    
    start = end + 1
    end += len(fr2)

    format_cells(wb, ws, start, end, 'B8', 'G13', col_size)
    
    start = end + 1
    end += len(so1)

    format_cells(wb, ws, start, end, 'B14', 'G18', col_size)
    
    start = end + 1
    end += len(so2)

    format_cells(wb, ws, start, end, 'B19', 'G23', col_size)
    
    start = end + 1
    end += len(jr1)

    format_cells(wb, ws, start, end, 'B24', 'G29', col_size)
    
    start = end + 1
    end += len(jr2)

    format_cells(wb, ws, start, end, 'B30', 'G34', col_size)
    
    start = end + 1
    end += len(sr1)

    format_cells(wb, ws, start, end, 'B35', 'G40', col_size)
    
    start = end + 1
    end += len(sr2)

    format_cells(wb, ws, start, end, 'B41', 'G44', col_size)

    start = end + 1
    end += len(unused)

    end_cell = 'F' + str(end)

    ws.merge_cells(start_row = start, start_column = 1, end_row = end, end_column = 1)
    cells = list(ws['B45':end_cell])
    i = 0
    for row in cells:
        j = 0
        row_size = len(row)
        #find all courses that have not been taken yet
        for cell in row:
            #if first index, then we are at the top border cell
            if i == 0:
                #if first index, we are at the left border cell
                if j == 0:
                    cells[i][j].border = top_left
                #if last index, then we are at the right border cell
                elif j+1 == row_size:
                    cells[i][j].border = top_right
                else:
                    cells[i][j].border = top
            #if last index, then we are at the bottom border cell
            elif i+1 == col_size:
                if j == 0:
                    cells[i][j].border = bottom_left
                elif j+1 == row_size:
                    cells[i][j].border = bottom_right
                else:
                    cells[i][j].border = bottom
            elif j == 0:
                cells[i][j].border = left
            elif j+1 == row_size:
                cells[i][j].border = right
            #for inner cells, make border thin
            else:
                cells[i][j].border = other
            j += 1
        i += 1
        
    #save changes to excel file
    wb.save(path)
