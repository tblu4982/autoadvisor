import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import re

class configure_file():
    def get_struct(self):
        return self.courses_struct

    def get_courses(self):
        return self.core_course_stack, self.elec_stack

    def get_sems(self):
        return self.frs1, self.frs2, self.sos1, self.sos2, \
               self.jrs1, self.jrs2, self.srs1, self.srs2

    def set_config(self, config_file):
        proto = []

        #get course structure from excel spreadsheet
        try:
            wb = load_workbook(self.config_file)    
            name = wb.sheetnames[0]
            sheet = wb[name]

            in_sem = False
            sem = ""
            for row in sheet.rows:
                special_case = False
                row = list(row)
                course_type = row[1].value
                course_id = row[2].value
                course_name = row[3].value
                min_grade = row[4].value
                min_credits = row[6].value
                #skip rows until we reach first semester
                if not in_sem:
                    if not row[0].value == None:
                        if re.search("Freshman S1", row[0].value):
                            in_sem = True
                if in_sem:
                    if not row[0].value == None:
                        sem = row[0].value
                    #put core courses into core course stack
                    if re.search("core", course_type):
                        self.core_course_stack[course_id] = [course_id, course_name, '', '', '', '']
                    #put all other courses into elec stack
                    else:
                        self.elec_stack[course_id] = [course_id, course_name, '', '', '', '']
                    #structure semester stacks
                    if re.search('Freshman S1', sem):
                        self.frs1[course_id] = [course_id, course_name, '', '', '', '']
                    elif re.search('Freshman S2', sem):
                        self.frs2[course_id] = [course_id, course_name, '', '', '', '']
                    elif re.search('Sophomore S1', sem):
                        self.sos1[course_id] = [course_id, course_name, '', '', '', '']
                    elif re.search('Sophomore S2', sem):
                        self.sos2[course_id] = [course_id, course_name, '', '', '', '']
                    elif re.search('Junior S1', sem):
                        self.jrs1[course_id] = [course_id, course_name, '', '', '', '']
                    elif re.search('Junior S2', sem):
                        self.jrs2[course_id] = [course_id, course_name, '', '', '', '']
                    elif re.search('Senior S1', sem):
                        self.srs1[course_id] = [course_id, course_name, '', '', '', '']
                    elif re.search('Senior S2', sem):
                        self.srs2[course_id] = [course_id, course_name, '', '', '', '']
                    #structure courses_struct for course requirements
                    if not course_type == None:
                        proto.append(course_type)
                    if not min_grade == None:
                        proto.append(min_grade)
                    if not min_credits == None:
                        proto.append(min_credits)
                    else:
                        proto.append('')
                    for i in range(7, 12):
                        if not row[i].value == None:
                            proto.append(row[i].value)
                        else:
                            proto.append('')
                    #check if course has a prereq, coreq, a sub, or eligible courses
                    for i in proto:
                        if not i == '':
                            special_case = True
                            break
                    #if course is a special case, we need to mark it in course struct
                    if special_case:
                        course = proto.copy()
                        self.courses_struct[course_id] = course
                    proto.clear()
            wb.close()
        except InvalidFileException:
            pass

    def __init__(self, config):
        self.config_file = config
        #core_course_stack structure:
        #ID, Name, Grade, Credits, Semester, Notes
        self.core_course_stack = {}
        #elec_stack structure:
        #ID, Name, Grade, Credits, Semester, Notes
        self.elec_stack = {}
        #courses_struct structure:
        #Course Type, Min Grade, Min Credits, Prereq, Coreq, Subs, Eligible Courses, Eligible Course Prereqs
        self.courses_struct = {}
        #semester structures:
        self.frs1 = {}
        self.frs2 = {}
        self.sos1 = {}
        self.sos2 = {}
        self.jrs1 = {}
        self.jrs2 = {}
        self.srs1 = {}
        self.srs2 = {}
        self.set_config(self.config_file)
