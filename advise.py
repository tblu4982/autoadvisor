import re #used for regular expressions
from datetime import datetime #use to get current date
import pandas as pd #used to structure data
import openpyxl #used to create excel files
from openpyxl.workbook import Workbook #used to generate worksheets in memory
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side #used for excel border styles
from openpyxl.styles import PatternFill #used for excell cell background colors
from openpyxl.styles.colors import Color
from openpyxl.styles import Alignment
import open_course_struct
import sys

#global variables for border styles
top = Border(left = Side(style = 'thin'),
              right = Side(style = 'thin'),
              top = Side(style = 'thick'),
              bottom = Side(style = 'thin'))
top_left = Border(left = Side(style = 'thick'),
              right = Side(style = 'thin'),
              top = Side(style = 'thick'),
              bottom = Side(style = 'thin'))
top_right = Border(left = Side(style = 'thin'),
              right = Side(style = 'thick'),
              top = Side(style = 'thick'),
              bottom = Side(style = 'thin'))
right = Border(left = Side(style = 'thin'),
              right = Side(style = 'thick'),
              top = Side(style = 'thin'),
              bottom = Side(style = 'thin'))
bottom_right = Border(left = Side(style = 'thin'),
              right = Side(style = 'thick'),
              top = Side(style = 'thin'),
              bottom = Side(style = 'thick'))
bottom = Border(left = Side(style = 'thin'),
              right = Side(style = 'thin'),
              top = Side(style = 'thin'),
              bottom = Side(style = 'thick'))
bottom_left = Border(left = Side(style = 'thick'),
              right = Side(style = 'thin'),
              top = Side(style = 'thin'),
              bottom = Side(style = 'thick'))
left = Border(left = Side(style = 'thick'),
              right = Side(style = 'thin'),
              top = Side(style = 'thin'),
              bottom = Side(style = 'thin'))
other = Border(left = Side(style = 'thin'),
              right = Side(style = 'thin'),
              top = Side(style = 'thin'),
              bottom = Side(style = 'thin'))
#global variables for cell background colors
redFill = PatternFill('solid', fgColor = 'FF0000')
greenFill = PatternFill('solid', fgColor = '00FF00')
yellowFill = PatternFill('solid', fgColor = 'FFFF00')
blueFill = PatternFill('solid', fgColor = '00BFFF')
orangeFill = PatternFill('solid', fgColor = 'FCB900')

#at least 8 csci courses, 60 credits, and csci 287 before 400-level csci courses
#method to check if a course has a passing grade where a passing grade is 'C'
def passing_grade(grade, sem, curr_sem):
    if grade <= "C" or grade == "S" or grade == "TR" or grade == "In progress":
        return True
    #If we reached this, then we reached a failing course somehow
    else:
        return False

#Check if course satisfies elective
def check_course(course, key):
    #pulls list of courses from config file that can satisfy elective
    valid_courses = courses_struct[key][6].split(',')
    for i in range(len(valid_courses)):
        valid_courses[i] = valid_courses[i].strip()
    for c in valid_courses:
        #if the course exists within config file, then we have a match
        if re.search(c, course):
            return True
    return False
    
#Finds the current semester
def get_curr_sem():
    month = datetime.now().month
    year = datetime.now().year
    #returns the correct semester based on datetime values
    if month >= 1:
        if month >= 5:
            if month <= 8:
                if month == 12:
                    return "WI" + str(year)[-2:]
                else:
                    return "FA" + str(year)[-2:]
            else:
                return "SU" + str(year)[-2:]
        else:
            return "SP" + str(year)[-2:]

#checks if course has valid subs
def subs_check(courses, course, unused):
    #pulls list of courses that can be used as a sub
    subs = courses_struct[course][5].split(',')
    #store the keys of all courses in the 'unused' dictionary
    keys = unused.keys()
    pop_key = ''
    found_sub = False
    #if there are subs, find them from unused
    if len(subs) > 0 and not subs[0] == '':
        for i in range(len(subs)):
            subs[i] = subs[i].strip()
            #compare the subs to all courses in 'unused'
            for key in keys:
                #set the missing course to the sub if a match is found
                if re.search(subs[i], key):
                    found_sub = True
                    c_id = courses[course][0]
                    courses[course] = unused[key]
                    courses[course][0] = c_id
                    courses[course].append(unused[key][0] + ' sub')
                    print(c_id)
                    print(unused[key][0])
                    pop_key = key
    #remove sub from unused
    if found_sub:
        unused.pop(pop_key)

#checks if student is eleigible for course
def eligibility_check(key, courses, total_credits, core_courses):
    course_type = courses_struct[key][0]
    try:
        courses[key][5] = ''
    except IndexError:
        courses[key].append("")
    #get prerequisites and put them in an array
    prereqs = courses_struct[key][3].split(',')
    min_credits = courses_struct[key][2]
    is_satisfied = True
    #counter for courses in which the student needs to take a certain
    #number of core courses
    count = 0
    c_count = 0
    #eligibility code that tells student why they aren't eligible for course
    elig_code = 0
    pop_prereqs = []
    #if there are prerequisites, check if they have been taken
    if bool(prereqs[0]):
        for i in range(len(prereqs)):
            prereqs[i] = prereqs[i].strip()
            if not re.search('\*', prereqs[i]):
                try:
                    if not bool(core_courses[prereqs[i]][2]):
                        is_satisfied = False
                        if courses[key][2] == 'In progress':
                            elig_code = 3
                        else:
                            elig_code = 1
                    else:
                        pop_prereqs.append(i)
                except KeyError:
                    break
            #for prereqs in which the student has to take 'x' number of courses
            else:
                #search criteria and number of courses
                c_search, c_count = prereqs[i].split('*')
                c_count = int(c_count)
                #search through courses for courses that match criteria
                for course in courses:
                    #count only courses that have been passed and are not labs
                    if re.search(c_search, course) and bool(courses[course][2]):
                        if not courses[course][2] == 'In progress':
                            if not float(courses[course][3]) == 1:
                                count += 1
                #if not enough courses for criteria, print why
                if count < c_count:
                    is_satisfied = False
                    elig_code = 2
                    
    #check if transcript meets credit requirements for course
    if bool(min_credits):
        if not total_credits >= courses_struct[key][2]:
            is_satisfied = False
            credit_diff = courses_struct[key][2] - total_credits
            courses[key][5] = str(credit_diff) + ' core credits short, '

    #remove satisfied prerequisites
    pop_prereqs.reverse()
    for i in pop_prereqs:
        prereqs.pop(i)

    #ERROR: #V00658286, LIT ELECTIVE NOT BEING LOCKED WHEN ENGL 111 IS MISSING
    if is_satisfied:
        #check if course is an elective
        if course_type == 'elective':
            #list of courses that can satisfy elective
            valid_courses = courses_struct[key][6].split(',')
            #list of prerequisites for electives that require them
            elig_courses = courses_struct[key][7].split(',')
            #holds all courses whose prereqs are satisfied
            e_courses = []
            #holds all courses whose prereqs are not satisfied
            ineligible = []
            #check if any of the courses that can satisfy elective
            #are already on transcript
            for i in range(len(valid_courses)):
                is_valid = True
                valid_courses[i] = valid_courses[i].strip()
                #special case: science lectures and labs on transcript share the
                #same course ID, differentiate between them to avoid the issue
                #lectures being counted as labs and vice versa
                for course in courses:
                    if re.search('^SCI', key) and re.search('^SCI', course):
                        if not re.search('LAB$', key):
                            if not re.search('LAB$', course):
                                #If any of the courses that can satisfy elective
                                #are found, do not include it in recommendations
                                if valid_courses[i] == courses[course][0]:
                                    is_valid = False
                        else:
                            if re.search('LAB$', course):
                                if valid_courses[i] == courses[course][0]:
                                    is_valid = False
                if is_valid:
                    #if none of the suitable courses have prereqs, then student
                    #is eligible to take them all
                    if not elig_courses[0] == '':
                        #scan through prerequisites
                        for course in elig_courses:
                            #assign the target course and the required prereq
                            try:
                                e_course, e_prereqs = course.strip().split(':')
                            except ValueError:
                                sys.exit('Please check configuration file! Eligible' \
                                         + ' course prerequisites for row ' + key + ' is' \
                                         + ' not properly structured! (Error occurred after ' \
                                         + e_course + ')')
                            e_course = e_course[1:]
                            #ISSUE WITH INCORRECT COURSES BEING RECOMMENDED
                            #MAY BE HERE
                            e_prereqs = e_prereqs.split('&')
                            j = 0
                            #format prerequisites by stripping whitespace and removing tags
                            for e_prereq in e_prereqs:
                                e_prereqs[j] = e_prereq.strip()
                                if re.search('>$', e_prereqs[j]):
                                    e_prereqs[j] = e_prereqs[j][:-1]
                                j += 1
                            #if target course matches current index of valid courses,
                            #search to see if prereq is satisfied
                            if valid_courses[i] == e_course:
                                for e_prereq in e_prereqs:
                                    try:
                                        if core_courses[e_prereq][2] == '':
                                            #if prereq not found, then student not eligible for course
                                            is_valid = False
                                            #append prereq to list 
                                            if not e_prereq in ineligible:
                                                ineligible.append(e_prereq)
                                    #if this error occurs, then try reading through elec stack
                                    except KeyError:
                                        try:
                                            is_found = False
                                            for elec_course in courses:
                                                #if prereq is found in elec stack, break loop
                                                if courses[elec_course][0] == e_prereq:
                                                    if not courses[elec_course][2] == '':
                                                        is_found = True
                                                        break
                                            #if prereq is not found
                                            if not is_found:
                                                is_valid = False
                                                if not e_prereq in ineligible:
                                                    ineligible.append(e_prereq)
                                        #break loop if this error occurs
                                        except IndexError:
                                            break
                #if all prerequisites are satisfied, student is eligible for this elective course
                if is_valid:
                    e_courses.append(valid_courses[i])
                    
            #if student not eligible for elective, print required courses
            if len(e_courses) == 0:
                if courses[key][2] == "In progress":
                    courses[key][5] += 'Drop course! '
                courses[key][5] += 'Need '
                for c in ineligible:
                    courses[key][5] += c + ', '
                return
            #if student is eligible, print eligible courses that can satisfy it
            #omit courses in progress
            elif not courses[key][2] == 'In progress':
                courses[key][5] += 'Eligible courses: '
                for c in e_courses:
                    if len(c) < 8:
                        courses[key][5] += 'Any ' + c + ' course, '
                    else:
                        courses[key][5] += c + ', '
                return

    #if course is satisfied, print eligibility
    if is_satisfied and elig_code == 0 and not courses[key][2] == 'In progress':
        courses[key][5] = 'Eligible'
    #if we reach this, there are courses that need to be taken first
    if elig_code == 1:
        courses[key][5] += 'Need '
        for prereq in prereqs:
            courses[key][5] += prereq + ', '
    #if we reach this, student lacks number of courses needed to satisfy criteria
    if elig_code == 2:
        courses[key][5] += 'Need '
        for prereq in prereqs:
            if re.search('\*', prereq):
                course = prereq.split('*')[0]
                courses[key][5] += str(c_count - count) + ' ' + course + ' courses, '
            else:
                courses[key][5] += prereq + ', '
    #if we reach this, student is enrolled in a course that they lack prereqs for
    if elig_code == 3:
        courses[key][5] += 'Drop course! Need '
        for prereq in prereqs:
            courses[key][5] += prereq + ', '

#Finds common courses between dictionaries
def find_courses(semester, core_courses, electives):
    sem_keys = semester.keys()
    core_keys = core_courses.keys()
    elec_keys = electives.keys()

    for key in sem_keys:
        if key in core_keys:
            if bool(core_courses[key]):
                semester[key] = core_courses[key]
                
    for key in sem_keys:
        if key in elec_keys:
            if bool(electives[key]):
                semester[key] = electives[key]

#method to add 6th index to elements missing it
def structure_dict_elems(semester):
    keys = semester.keys()
    for key in semester:
        #Some elements already have 6 indices
        #check for length of course
        #if course length is 6, do not append
        if len(semester[key]) < 6:
            semester[key].append('')
        #if course length is somehow greater than 6, pop indices
        if len(semester[key]) > 6:
            semester[key].pop()

#map elective keys from planning sheet to elective keys in config file
def map_elec(course):
    if course == 'HISTELEC1':
        return 'HIST'
    if course == 'SOCELEC1':
        return 'SS'
    if course == 'GLOELEC1':
        return 'GS'
    if course == 'LITELEC1':
        return 'LIT'
    if course == 'SCILEC1':
        return 'SCI 1'
    if course == 'SCILAB1':
        return 'SCI 1 LAB'
    if course == 'FREEELEC1':
        return 'FREE 1'
    if course == 'CSCIELEC3':
        return 'CMS'
    if course == 'FREEELEC2':
        return 'FREE 2'
    if course == 'CSCIELEC1':
        return 'CS 1'
    if course == 'SCILEC2':
        return 'SCI 2'
    if course == 'SCILAB2':
        return 'SCI 2 LAB'
    if course == 'CSCIELEC2':
        return 'CS 2'
    if course == 'MATHELEC1':
        return 'MATH'

#determines which color the cells will be colored
def find_fill_type(is_complete, in_progress, is_open, is_valid):
    #for all taken or in progress courses, mark green and yellow respectively
    if is_complete:
        #for courses in progress, mark yellow
        if in_progress:
            if is_valid:
                return yellowFill
            else:
                return orangeFill
        #for completed courses, mark green
        else:
            return greenFill
    #For courses the student is eligible for, mark blue
    elif is_open:
        return blueFill
    #For all other courses, mark red
    else:
        return redFill
    

#Merges index columns and styles the cells
def format_cells(wb, ws, start, end, start_cell, end_cell, col_size):
    #finds index columns and merges them
    ws.merge_cells(start_row = start, start_column = 1, end_row = end, end_column = 1)
    #grabs range of cells that align with merged indices, and creates a border
    cells = list(ws[start_cell:end_cell])
    i = 0
    for row in cells:
        j = 0
        row_size = len(row) - 1
        is_complete = True
        in_progress = False
        is_open = False
        is_valid = True
        #find all courses that either have not been taken yet or are in progress
        for cell in row:
            if cell.value == None:
                if not row[j] == row[-1]:
                    is_complete = False
            if cell.value == "In progress":
                in_progress = True
            try:
                if re.search('Eligible', cell.value):
                    is_open = True
                if re.search('Drop', cell.value):
                    is_valid = False
            except TypeError:
                continue
            j += 1
        j = 0
        for cell in row:
            #if first index, then we are at the top border cell
            if i == 0:
                #if first index, we are at the left border cell
                if j == 0:
                    cells[i][j].border = top_left
                    #find which color to make the cell
                    cells[i][j].fill = find_fill_type(is_complete, in_progress, is_open, is_valid)
                #if last index, then we are at the right border cell
                elif j+1 == row_size:
                    cells[i][j].border = top_right
                    cells[i][j].fill = find_fill_type(is_complete, in_progress, is_open, is_valid)
                elif not row[j] == row[-1]:
                    cells[i][j].border = top
                    cells[i][j].fill = find_fill_type(is_complete, in_progress, is_open, is_valid)
            #if last index, then we are at the bottom border cell
            elif i+1 == col_size:
                if j == 0:
                    cells[i][j].border = bottom_left
                    cells[i][j].fill = find_fill_type(is_complete, in_progress, is_open, is_valid)
                elif j+1 == row_size:
                    cells[i][j].border = bottom_right
                    cells[i][j].fill = find_fill_type(is_complete, in_progress, is_open, is_valid)
                elif not row[j] == row[-1]:
                    cells[i][j].border = bottom
                    cells[i][j].fill = find_fill_type(is_complete, in_progress, is_open, is_valid)
            elif j == 0:
                cells[i][j].border = left
                cells[i][j].fill = find_fill_type(is_complete, in_progress, is_open, is_valid)
            elif j+1 == row_size:
                cells[i][j].border = right
                cells[i][j].fill = find_fill_type(is_complete, in_progress, is_open, is_valid)
            #for inner cells, make border thin
            elif not row[j] == row[-1]:
                cells[i][j].border = other
                cells[i][j].fill = find_fill_type(is_complete, in_progress, is_open, is_valid)
            j += 1
            if j == len(row):
                break
        i += 1

def main(courses, name, config_file, fullname, vnum, advisor):
    print("Generating Advisory Report for " + name + "...")
    print(vnum)

    #set core courses, electives, and courses_struct from configuration file
    global courses_struct
    configuration = open_course_struct.configure_file(config_file)
    courses_struct = configuration.get_struct()

    core_course_stack, elec_stack = configuration.get_courses()

    #variable to hold the current semester
    curr_sem = get_curr_sem()

    #Get credits from transcript
    total_credits = 0
    
    unused = {}
    i = 0
    j = 0

    #loop through matrix to find courses matching requirements
    for c in courses:
        passed_course = False
        course = c[0]
        cname = c[1]
        grade = c[2]
        credit = float(c[3])
        term = c[4]
        abbr = course[:5]
        #Global Studies and History Electives have some overlap
        #If 'HIST 114' and 'HIST 115' are encountered, save them for later
        #If Global Studies and History aren't satisfied, fill Global Studies first
        if course == 'HIST 114' or course == 'HIST 115':
            try:
                #if duplicate courses are found, give them a unique key
                #otherwise, old course will be replaced with new course
                if bool(unused[course]):
                    course = course + 'd' + str(i)
                    unused[course] = c
                    i += 1
            #if key doesn't exist in unused, then assign course to that key
            except KeyError:
                unused[course] = c
            continue
        #search for core class requirements first
        if course in core_course_stack:
            min_grade = courses_struct[course][1]
            #for certain courses, 'C' is a passing grade
            if min_grade == 'C':
                passed_course = passing_grade(grade, term, curr_sem)
            else:
                passed_course = True
            if passed_course:
                total_credits += credit
                core_course_stack[course] = c
        #for non-core courses
        else:
            #assume course doesn't satisfy any elective, and prove this to be false
            is_found = False
            keys = []
            #set keys to only search through unsatisfied electives
            for key in elec_stack:
                if not bool(elec_stack[key][2]):
                    keys.append(key)
            #search through unsatisfied electives to see if course matches
            for key in keys:
                elec_course = elec_stack[key][2]
                course_type = courses_struct[key][0]
                #if elective has no grade, then it is likely not taken
                if not bool(elec_course):
                    #check if course in stack is elective or free
                    if course_type == 'elective':
                        #check if current course satisfies list of eligible courses
                        if check_course(course, key):
                            #this course satisfies an elective
                            is_found = True
                            if re.search('^SCI ', key):
                                #check if key is lab or not
                                if not re.search('LAB$', key):
                                    #sci courses with 4 credits satisfy both course and lab
                                    if credit == 4:
                                        #set the corresponding lab to the lecture course
                                        new_struct = c.copy()
                                        elec_stack[key][0] = c[0]
                                        elec_stack[key][5] = elec_stack[key][1] + ' satisfied'
                                        elec_stack[key][1] = c[1]
                                        elec_stack[key][2] = c[2]
                                        elec_stack[key][3] = c[3]
                                        elec_stack[key][4] = c[4]
                                        new_key = key + ' LAB'
                                        new_struct[2] = 'N/A'
                                        new_struct[3] = 'N/A'
                                        elec_stack[new_key][0] = new_struct[0]
                                        elec_stack[new_key][5] = elec_stack[new_key][1] + ' satisfied'
                                        elec_stack[new_key][1] = new_struct[1]
                                        elec_stack[new_key][2] = new_struct[2]
                                        elec_stack[new_key][3] = new_struct[3]
                                        elec_stack[new_key][4] = new_struct[4]
                                        total_credits += credit
                                        break
                                    #sci courses with 3 credits are lecture courses
                                    elif credit == 3:
                                        elec_stack[key][0] = c[0]
                                        elec_stack[key][5] = elec_stack[key][1] + ' satisfied'
                                        elec_stack[key][1] = c[1]
                                        elec_stack[key][2] = c[2]
                                        elec_stack[key][3] = c[3]
                                        elec_stack[key][4] = c[4]
                                        total_credits += credit
                                        break
                                    #if lab course is listed before lecture, set key to
                                    #lab and assign the course there
                                    #possible error: If lab is fulfilled but lecture is not
                                    #                next time we encounter a lab, it may replace
                                    #                old lab course
                                    elif credit == 1:
                                        elec_stack[key + ' LAB'][0] = c[0]
                                        elec_stack[key + ' LAB'][5] = elec_stack[key + ' LAB'][1] + ' satisfied'
                                        elec_stack[key + ' LAB'][1] = c[1]
                                        elec_stack[key + ' LAB'][2] = c[2]
                                        elec_stack[key + ' LAB'][3] = c[3]
                                        elec_stack[key + ' LAB'][4] = c[4]
                                        total_credits += credit
                                        break
                                #sci courses with 1 credit is a lab course
                                elif credit == 1:
                                    elec_stack[key][0] = c[0]
                                    elec_stack[key][5] = elec_stack[key][1] + ' satisfied'
                                    elec_stack[key][1] = c[1]
                                    elec_stack[key][2] = c[2]
                                    elec_stack[key][3] = c[3]
                                    elec_stack[key][4] = c[4]
                                    total_credits += credit
                                    break
                            #for all other electives
                            else:
                                if courses_struct[key][1] == 'C':
                                    passed_course = passing_grade(grade, term, curr_sem)
                                else:
                                    passed_course = True

                                if passed_course:
                                    elec_stack[key][0] = c[0]
                                    elec_stack[key][5] = elec_stack[key][1] + ' satisfied'
                                    elec_stack[key][1] = c[1]
                                    elec_stack[key][2] = c[2]
                                    elec_stack[key][3] = c[3]
                                    elec_stack[key][4] = c[4]
                                    total_credits += credit
                                    break
            #store unused courses into unused stack
            if not is_found:
                try:
                    #if duplicate courses are found, give them a unique key
                    if bool(unused[course]):
                        course = course + 'd' + str(i)
                        unused[course] = c
                        i += 1
                except KeyError:
                    unused[course] = c

    #check unused stack to see if any course satisfies core course as a sub
    for course in core_course_stack:
        if core_course_stack[course][2] == '':
            subs_check(core_course_stack, course, unused)

    #check if global studies and/or history electives are unsatisfied, and
    #assign stored history courses if applicable
    pop_keys = []
    #Global studies takes higher priority, set as first index to search
    keys = ['GS', 'HIST']
    for course in unused:
        for key in keys:
            if not bool(elec_stack[key][2]):
                if check_course(course, key):
                    elec_stack[key][0] = unused[course][0]
                    elec_stack[key][5] = elec_stack[key][1] + ' satisfied'
                    elec_stack[key][1] = unused[course][1]
                    elec_stack[key][2] = unused[course][2]
                    elec_stack[key][3] = unused[course][3]
                    elec_stack[key][4] = unused[course][4]
                    pop_keys.append(course)
                    total_credits += credit
                    break

    for key in pop_keys:
        unused.pop(key)
            
    #fill free electives from unused courses
    pop_keys = []
    for course in unused:
        #exclude courses that don't have 3 credits
        if float(unused[course][3]) >= 3:
            #only search through electives that aren't satisfied
            for key in elec_stack:
                if not bool(elec_stack[key][2]):
                    elec_course = elec_stack[key][2]
                    course_type = courses_struct[key][0]
                    #check if elective is empty
                    if not bool(elec_course):
                        #check if course is a free elective
                        if course_type == 'free':
                            elec_stack[key][0] = unused[course][0]
                            elec_stack[key][5] = elec_stack[key][1] + ' satisfied'
                            elec_stack[key][1] = unused[course][1]
                            elec_stack[key][2] = unused[course][2]
                            elec_stack[key][3] = unused[course][3]
                            elec_stack[key][4] = unused[course][4]
                            pop_keys.append(course)
                            total_credits += credit
                            break

    for key in pop_keys:
        unused.pop(key)

    #add a note to core courses that the student is eligible to take
    for course in core_course_stack:
        #Student may be enrolled in courses they aren't eligible for, check for those as well
        if not bool(core_course_stack[course][2]) or core_course_stack[course][2] == "In progress":
            eligibility_check(course, core_course_stack, total_credits, core_course_stack)
            

    #add a note to electives that the student is eligible to take
    for course in elec_stack:
        if not bool(elec_stack[course][2]) or elec_stack[course][2] == "In progress":
            eligibility_check(course, elec_stack, total_credits, core_course_stack)

    #Set semester structures from configuration file
    frs1, frs2, sos1, sos2, jrs1, jrs2, srs1, srs2 = configuration.get_sems()

    #assign courses to semesters
    find_courses(frs1, core_course_stack, elec_stack)
    find_courses(frs2, core_course_stack, elec_stack)
    find_courses(sos1, core_course_stack, elec_stack)
    find_courses(sos2, core_course_stack, elec_stack)
    find_courses(jrs1, core_course_stack, elec_stack)
    find_courses(jrs2, core_course_stack, elec_stack)
    find_courses(srs1, core_course_stack, elec_stack)
    find_courses(srs2, core_course_stack, elec_stack)
    
    #Structure dictionaries to add/remove indices to match desired structure
    structure_dict_elems(frs1)
    structure_dict_elems(frs2)
    structure_dict_elems(sos1)
    structure_dict_elems(sos2)
    structure_dict_elems(jrs1)
    structure_dict_elems(jrs2)
    structure_dict_elems(srs1)
    structure_dict_elems(srs2)

    #create index column for excel file
    index = []
    for i in range(len(frs1)):
        index.append('Freshman: Semester 1')
    for i in range(len(frs2)):
        index.append('Freshman: Semester 2')
    for i in range(len(sos1)):
        index.append('Sophomore: Semester 1')
    for i in range(len(sos2)):
        index.append('Sophomore: Semester 2')
    for i in range(len(jrs1)):
        index.append('Junior: Semester 1')
    for i in range(len(jrs2)):
        index.append('Junior: Semester 2')
    for i in range(len(srs1)):
        index.append('Senior: Semester 1')
    for i in range(len(srs2)):
        index.append('Senior: Semester 2')
    for i in range(len(unused)):
        index.append('Unused Courses')

    #merge semesters into one list
    courses.clear()
    fr1 = list(frs1.values())
    for course in fr1:
        courses.append(course)
    fr2 = list(frs2.values())
    for course in fr2:
        courses.append(course)
    so1 = list(sos1.values())
    for course in so1:
        courses.append(course)
    so2 = list(sos2.values())
    for course in so2:
        courses.append(course)
    jr1 = list(jrs1.values())
    for course in jr1:
        courses.append(course)
    jr2 = list(jrs2.values())
    for course in jr2:
        courses.append(course)
    sr1 = list(srs1.values())
    for course in sr1:
        courses.append(course)
    sr2 = list(srs2.values())
    for course in sr2:
        courses.append(course)
    unused_c = list(unused.values())
    for course in unused_c:
        courses.append(course)

    #set excel file path
    path = 'advisors\\' + advisor + '\\' + name + '\\' + config_file.split('/')[-1].split('.')[0] + '\\' + name + '_planning_sheet(Student).xlsx'

    #structure index and coursesd into a dataframe
    df = pd.DataFrame(courses,
                      index,
                      columns = ['ID', 'Name', 'Grade', 'Credits', 'Semester', 'Notes'])

    #sheet name must not exceed 31 characters
    if len(name + "(" + vnum + ")") > 31:
        df.to_excel(path, sheet_name = "(" + vnum + ")")
    else:
        df.to_excel(path, sheet_name = name + "(" + vnum + ")")
    #modify created excel file
    wb = load_workbook(filename = path)
    ws = wb.active
    #set column width for spreadsheet
    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 7
    ws.column_dimensions['F'].width = 12
    #center text for certain columns
    for row in ws:
        cells = [row[0], row[3], row[5]]
        for cell in cells:
            if cell == cells[0]:
                cell.alignment = Alignment(wrap_text = True, horizontal = 'center', vertical = 'center')
            else:
                cell.alignment = Alignment(horizontal = 'center')
    #merge cells in first colum to remove redundant information
    col_size = len(courses)

    start = 2
    end = len(frs1) + 1
    upper_range = 'B' + str(start)
    lower_range = 'G' + str(end)
    #color code rows
    format_cells(wb, ws, start, end, upper_range, lower_range, col_size)
    
    start = end + 1
    end += len(frs2)

    upper_range = 'B' + str(start)
    lower_range = 'G' + str(end)

    format_cells(wb, ws, start, end, upper_range, lower_range, col_size)
    
    start = end + 1
    end += len(sos1)

    upper_range = 'B' + str(start)
    lower_range = 'G' + str(end)

    format_cells(wb, ws, start, end, upper_range, lower_range, col_size)
    
    start = end + 1
    end += len(sos2)

    upper_range = 'B' + str(start)
    lower_range = 'G' + str(end)

    format_cells(wb, ws, start, end, upper_range, lower_range, col_size)
    
    start = end + 1
    end += len(jrs1)

    upper_range = 'B' + str(start)
    lower_range = 'G' + str(end)

    format_cells(wb, ws, start, end, upper_range, lower_range, col_size)
    
    start = end + 1
    end += len(jrs2)

    upper_range = 'B' + str(start)
    lower_range = 'G' + str(end)

    format_cells(wb, ws, start, end, upper_range, lower_range, col_size)
    
    start = end + 1
    end += len(srs1)

    upper_range = 'B' + str(start)
    lower_range = 'G' + str(end)

    format_cells(wb, ws, start, end, upper_range, lower_range, col_size)
    
    start = end + 1
    end += len(srs2)

    upper_range = 'B' + str(start)
    lower_range = 'G' + str(end)

    format_cells(wb, ws, start, end, upper_range, lower_range, col_size)

    if len(unused) > 0:
        start = end + 1
        end += len(unused)

        #unused courses do not need to be color coded, also have a smaller range
        upper_range = 'B' + str(start)
        lower_range = 'F' + str(end)

        ws.merge_cells(start_row = start, start_column = 1, end_row = end, end_column = 1)
        cells = list(ws[upper_range:lower_range])
        i = 0
        for row in cells:
            j = 0
            row_size = len(row)
            #find all courses that have not been taken yet
            for cell in row:
                #if first index, then we are at the top border cell
                if i == 0:
                    #if first index, we are at the left border cell
                    if j == 0:
                        cells[i][j].border = top_left
                    #if last index, then we are at the right border cell
                    elif j+1 == row_size:
                        cells[i][j].border = top_right
                    else:
                        cells[i][j].border = top
                #if last index, then we are at the bottom border cell
                elif i+1 == col_size:
                    if j == 0:
                        cells[i][j].border = bottom_left
                    elif j+1 == row_size:
                        cells[i][j].border = bottom_right
                    else:
                        cells[i][j].border = bottom
                elif j == 0:
                    cells[i][j].border = left
                elif j+1 == row_size:
                    cells[i][j].border = right
                #for inner cells, make border thin
                else:
                    cells[i][j].border = other
                j += 1
            i += 1

    start = end + 2
    upper_range = 'A' + str(start)
    cell = ws[upper_range]
    cell.value = "Please Read this!"

    upper_range = 'A' + str(start + 1)
    cell = ws[upper_range]
    cell.value = "For Science courses, student must take corresponding lab with lecture course!"

    upper_range = 'A' + str(start + 2)
    cell = ws[upper_range]
    cell.value = "For example, if you register for BIOL 120 with 3.0 credits (lecture), you need to register for BIOL 120 with 1.0 credit (lab)."
    
    upper_range = 'A' + str(start+3)
    cell = ws[upper_range]
    cell.value = "Make sure to take matching section number when signing up for Programming lecture and lab!"

    upper_range = 'A' + str(start+4)
    cell = ws[upper_range]
    cell.value = "For example, if you register for lecture CSCI 150-03, then register for Lab CSCI 151-03"
        
    #save changes to excel file
    wb.save(path)
    wb.close()

    

    path = 'advisors\\' + advisor + '\\' + name + '\\' + config_file.split('/')[-1].split('.')[0] + '\\' + name + '_planning_sheet(Advisor).xlsx'

    #open excel file
    wb = load_workbook(filename = 'Template Planning Sheet.xlsx')
    sheet_name = wb.sheetnames[2]
    sheet = wb[sheet_name]

    reached_sems = False

    #fill in courses to planning sheet
    i = 0
    for row in sheet.rows:
        #if we reach this, then we are at unused slot in planning sheet
        if i > 52:
            try:
                #--TO DO: Add code that adds rows to unused in planning sheet--
                #if unusued courses list is less than 12, all courses will fit
                if len(unused) <= 11:
                    keys = list(unused.keys())
                    #Course ID
                    row[2].value = unused[keys[0]][0]
                    #Course Title
                    row[3].value = unused[keys[0]][1]
                    #Course Credits
                    row[5].value = int(float(unused[keys[0]][3]))
                    #fill in course with 'in progress' or 'next sem' based on current semester
                    if unused[keys[0]][2] == 'In progress':
                        #if course is from current semester, add 'in progress'
                        if unused[keys[0]][4] == curr_sem:
                            row[6].value = 'In Progress'
                        #if course is not from current semester, add 'next sem'
                        else:
                            row[6].value = 'Next Sem'
                        row[7].value = unused[keys[0]][4]
                    else:
                        #Course Status
                        row[6].value = 'Completed'
                        #Course Term
                        row[7].value = unused[keys[0]][4]
                        #Course Grade
                        row[8].value = unused[keys[0]][2]
                    unused.pop(keys[0])
            #if this error occurs, skip it
            except AttributeError:
                i += 1
                continue
            #if this error occurs, stop loop
            except IndexError:
                break

        #skip rows in spreadsheet until we reach courses        
        if not reached_sems:
            if row[0].value == 'Semester, Year':
                reached_sems = True
        else:
            #identify course from planning sheet
            if not row[1].value == None:
                course = row[1].value
                #identify course by ID in planning sheet
                if re.search('[A-Z]{4}[0-9]{3}', course):
                    if len(course) > 7:
                        #for courses with IDs greater than 7, find exact course ID
                        course = re.search('[A-Z]{4}[0-9]{3}', course).group()
                    course = course[:-3] + ' ' + course[4:]
                #If Course ID can't be found in 1st column, check next column
                elif not row[2].value == None:
                    course = row[2].value
                #key elements in planning sheet:
                #HISTELEC1, SOCELEC1, GLOELEC1, LITELEC1, SCILEC1, SCILAB1
                #FREEELEC1, CSCIELEC3, FREEELEC2, CSCIELEC1, SCILEC2, SCILAB2
                #CSCIELEC2, MATHELEC1       
                try:
                    #get course from planning sheet, and find course from semester stacks
                    #variable to hold semester stack that target course is in
                    target_sem = ''
                    #map elective course key from planning sheet to config file elec key
                    if not re.search(' ', course):
                        course = map_elec(course)
                    if course in frs1:
                        target_sem = frs1
                    elif course in frs2:
                        target_sem = frs2
                    elif course in sos1:
                        target_sem = sos1
                    elif course in sos2:
                        target_sem = sos2
                    elif course in jrs1:
                        target_sem = jrs1
                    elif course in jrs2:
                        target_sem = jrs2
                    elif course in srs1:
                        target_sem = srs1
                    elif course in srs2:
                        target_sem = srs2
                    try:
                        #figure out why this is important
                        special = ['ENGL 342', 'PHIL 450']
                        if target_sem[course][2] == 'In progress':
                            if course in elec_stack or course in special:
                                row[2].value = target_sem[course][0]
                                row[4].value = target_sem[course][1]
                            if target_sem[course][4] == curr_sem:
                                row[6].value = 'In Progress'
                            else:
                                row[6].value = 'Next Sem'
                            row[7].value = target_sem[course][4]
                            if not re.search("satisfied$", target_sem[course][5]) and not target_sem[course][5] == "Eligible":
                                row[9].value = target_sem[course][5]
                        elif not target_sem[course][2] == '':
                            if course in elec_stack or course in special:
                                row[2].value = target_sem[course][0]
                                try:
                                    row[4].value = target_sem[course][1]
                                except AttributeError:
                                    pass
                            row[6].value = 'Completed'
                            row[7].value = target_sem[course][4]
                            row[8].value = target_sem[course][2]
                            if not re.search("satisfied$", target_sem[course][5]) and not target_sem[course][5] == "Eligible":
                                row[9].value = target_sem[course][5]
                        else:
                            if not re.search("satisfied$", target_sem[course][5]) and not target_sem[course][5] == "Eligible":
                                row[9].value = target_sem[course][5]
                    except TypeError:
                        i += 1
                        continue
                except AttributeError:
                    i += 1
                    continue
        i += 1

    wb.save(path)

    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]

    firstname = fullname[0]
    lastname = fullname[-1]

    for row in sheet.rows:
        proto = []
        if row[0].value == 'First Name:':
            row[1].value = firstname
        if row[0].value == 'Last Name:':
            row[1].value = lastname
        if row[0].value == 'V-number:':
            row[1].value = vnum

    wb.save(path)
    wb.close()
